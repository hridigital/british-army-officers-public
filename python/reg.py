#!/usr/bin/python3

#import sys
#import sqlite3
import re
from openpyxl import Workbook, load_workbook
from datetime import datetime
import traceback
#from tqdm import tqdm
from nameparser import HumanName
from reg_number import reg_number

ranks = {}

rwb = load_workbook(filename = 'data/lookups/Ranks.xlsx', read_only=True)
rws = rwb['Regiment']

for xli, xlrow in enumerate(rws.iter_rows()):
    for ci, cell in enumerate(xlrow):
        ranks[cell.value] = xli


#supp_ranks = {}
#supp_ranks['Solicitor'] = -1
#supp_ranks['Quarter-Master'] = -2
#supp_ranks['Surgeon'] = -3
#supp_ranks['Adjutant Surgeon'] = -4
#supp_ranks['Chaplain'] = -5
#supp_ranks['Chaplain General'] = -6

supp_ranks = {}

srwb = load_workbook(filename = 'data/lookups/RanksSupp.xlsx', read_only=True)
srws = srwb['Regiment']

for xli, xlrow in enumerate(srws.iter_rows()):
    for ci, cell in enumerate(xlrow):
        supp_ranks[cell.value] = xli



rank_lookups_regularised = {}
rank_lookups_unregularised = {}

rlwb = load_workbook(filename = 'data/lookups/RanksLookup.xlsx', read_only=True)
rlws = rlwb['Sheet1']

for xli, xlrow in enumerate(rlws.iter_rows()):
    if xlrow[2].value != None and len(xlrow[2].value) > 2:
        if xlrow[0].value != None and len(xlrow[0].value) > 2:
            rank_lookups_regularised[xlrow[0].value.strip()] = xlrow[2].value.strip()
        if xlrow[1].value != None and len(xlrow[1].value) > 2:
            rank_lookups_unregularised[xlrow[1].value.strip()] = xlrow[2].value.strip()


unit_lookups = {}

ulwb = load_workbook(filename = 'data/lookups/UnitsJuly24.xlsx', read_only=True)
ulws = ulwb['units']

for xli, xlrow in enumerate(ulws.iter_rows()):
    if xlrow[0].value != None and xlrow[3].value != None and len(xlrow[3].value) > 2:
        unit_lookups[xlrow[0].value.strip()[2:-2]] = xlrow[3].value.strip()




def rank_lvl(rank):
    if rank in supp_ranks:
        return -supp_ranks[rank]
    if rank in ranks:
        return ranks[rank]
    return None











def reg_regiment(regiment, year):

    if regiment == None: return None
    if isinstance(regiment, int): regiment = str(regiment)

    orig_regiment = regiment

    regiment = regiment.lower()
    regiment = re.sub(r'[^0-9a-z ]', r' ', regiment) # Replace any character which is not a lower case letter, a number or a space, with a space.
    regiment = re.sub(r' +', ' ', regiment) # Regularise space.
    regiment = ' ' + regiment + ' ' # Treat start or end of string as a word delimiter

    regiment = regiment.replace(' battallion', ' battalion')
    regiment = regiment.replace(' battlion', ' battalion')

    number = None
    unit = None

    if ' officers of the royal marines ' in regiment: unit = 'Royal Marines'; number = ''
    elif ' officers of the marine forces ' in regiment: unit = 'Royal Marines'; number = ''
    elif ' royal marine officers ' in regiment: unit = 'Royal Marines'; number = ''
    elif ' royal mar' in regiment: unit = 'Royal Marines'; number = ''
    elif ' r mar' in regiment: unit = 'Royal Marines'; number = ''

    elif ' royal engineers ' in regiment: unit = 'Royal Engineers'; number = ''

    elif ' royal regiment of horse guards ' in regiment: unit = 'Royal Horse Guards'; number = ''
    elif ' royal regiment of horse guard ' in regiment: unit = 'Royal Horse Guards'; number = ''
    elif ' r horse gds ' in regiment: unit = 'Royal Horse Guards'; number = ''
    elif ' r horse g ' in regiment: unit = 'Royal Horse Guards'; number = ''

    elif ' coldstream ' in regiment: unit = 'Coldstream Guards'; number = ''

    elif ' artillery ' in regiment: unit = 'Royal Artillery'; number = ''

    elif ' west indian rangers ' in regiment: unit = 'Royal West Indian Rangers'; number = ''

    elif ' west india regiment ' in regiment: unit = 'West India Regiment'

    elif ' life guards ' in regiment: unit = 'Life Guards'
    elif ' lifeguards ' in regiment: unit = 'Life Guards'
    elif ' life guard ' in regiment: unit = 'Life Guards'
    elif ' life gurds ' in regiment: unit = 'Life Guards'
    elif ' life gds ' in regiment: unit = 'Life Guards'
    elif ' life g ' in regiment: unit = 'Life Guards'



    elif ' rifle br' in regiment: unit = 'Rifle Brigade'; number = ''
    elif ' scotch brigade ' in regiment: unit = 'Scotch Brigade'; number = ''
    elif ' york hussars ' in regiment: unit = 'York Hussars'; number = ''
    elif " hompesch s mounted ri" in regiment: unit = "Hompesch's Mounted Riflemen"; number = ''

#    elif ' light battalion ' in regiment: unit = 'Light Battalion, KGL'
#    elif ' 2nd Light Battalion, KGL ' in regiment: unit = '2nd Light Battalion, KGL'

    elif ' king s german legion light infantry 'in regiment: unit = 'Light Battalion, KGL'
    elif ' kings german legion light infantry 'in regiment: unit = 'Light Battalion, KGL'

    elif ' the king s german legion infantry of the line 'in regiment: unit = 'Line Battalion, KGL'
    elif ' battalion of the line ' in regiment: unit = 'Line Battalion, KGL'
    elif ' bn of the line ' in regiment: unit = 'Line Battalion, KGL'
    elif ' line bn k g l ' in regiment: unit = 'Line Battalion, KGL'
    elif ' line k g l ' in regiment: unit = 'Line Battalion, KGL'
    elif ' bn line ' in regiment: unit = 'Line Battalion, KGL'

#    elif ' Garrison Company, KGL ' in regiment: unit = 'Garrison Company, KGL'
#    elif ' Veteran Battalion, KGL ' in regiment: unit = 'Veteran Battalion, KGL'
#    elif ' Artillery, KGL ' in regiment: unit = 'Artillery, KGL'
#    elif ' Engineers, KGL ' in regiment: unit = 'Engineers, KGL'
#    elif " Queen's German Regiment " in regiment: unit = "Queen's German Regiment"
#    elif " Duke of Brunswick Oel's Cavalry " in regiment: unit = "Duke of Brunswick Oel's Cavalry"
#    elif " Duke of Brunswick Oel's Infantry " in regiment: unit = "Duke of Brunswick Oel's Infantry"

    elif ' ceylon r' in regiment: unit = 'Ceylon Regiment'

#    elif ' 1st Greek Light Infantry ' in regiment: unit = '1st Greek Light Infantry'
#    elif ' 1st Greek Light Infantry ' in regiment: unit = '1st Greek Light Infantry'
#    elif " McDonnell's Foot " in regiment: unit = "McDonnell's Foot"
#    elif ' Loyal Liverpool Foot ' in regiment: unit = 'Loyal Liverpool Foot'
#    elif ' Loyal Sheffield Foot ' in regiment: unit = 'Loyal Sheffield Foot'
#    elif ' Londonderry Foot ' in regiment: unit = 'Londonderry Foot'
#    elif ' Royal Dublin Foot ' in regiment: unit = 'Royal Dublin Foot'
#    elif ' Royal Glasgow Foot ' in regiment: unit = 'Royal Glasgow Foot'
#    elif ' Loyal Cheshire Foot ' in regiment: unit = 'Loyal Cheshire Foot'
#    elif ' Loyal Kelso Foot ' in regiment: unit = 'Loyal Kelso Foot'
#    elif ' Exeter Volunteers ' in regiment: unit = 'Exeter Volunteers'
#    elif " Smyth's Foot " in regiment: unit = "Smyth's Foot"
#    elif ' Downshire Volunteers ' in regiment: unit = 'Downshire Volunteers'

    elif " armstrong " in regiment: unit = "Armstrong's Levy"; number = ''

#    elif " bissett " in regiment: unit = "Bissett's Levy"
#    elif " bradshaw " in regiment: unit = "Bradshaw's Levy"
#    elif " campbell " in regiment: unit = "Campbell's Levy"
#    elif ' french levy ' in regiment: unit = 'French Levy'
#    elif " hander " in regiment: unit = "Hander's Levy"
#    elif " loft " in regiment: unit = "Loft's Levy"
#    elif " kingston " in regiment: unit = "Kingston's Levy"
#    elif " nugent " in regiment: unit = "Nugent's Levy"
#    elif ' independent companies of invalids ' in regiment: unit = 'Independent Companies of Invalids'
#    elif ' corps of invalids in ireland ' in regiment: unit = 'Corps of Invalids in Ireland'

    elif ' royal garrison battalion ' in regiment: unit = 'Royal Garrison Battalion'; number = ''

    elif ' garrison battalion ' in regiment: unit = 'Garrison Battalion'

    elif ' royal veteran battalion' in regiment: unit = 'Royal Veteran Battalion'

#    elif ' New South Wales Veteran Company ' in regiment: unit = 'New South Wales Veteran Company'
#    elif ' European Garrison Companies ' in regiment: unit = 'European Garrison Companies'
#    elif ' Cape of Good Hope Veteran Company ' in regiment: unit = 'Cape of Good Hope Veteran Company'
#    elif ' Black Garrison Companies ' in regiment: unit = 'Black Garrison Companies'
#    elif ' Staff Garrison Companies ' in regiment: unit = 'Staff Garrison Companies'
#    elif ' New South Wales Corps ' in regiment: unit = 'New South Wales Corps'
#    elif " Queen's Rangers " in regiment: unit = "Queen's Rangers"
#    elif ' Canadian Fencibles ' in regiment: unit = 'Canadian Fencibles'
#    elif ' Canadian Voltigeurs ' in regiment: unit = 'Canadian Voltigeurs'
#    elif ' Glengarry Light Infantry ' in regiment: unit = 'Glengarry Light Infantry'
#    elif ' Michigan Fencibles ' in regiment: unit = 'Michigan Fencibles'
#    elif ' New Brunswick Provincial Regiment ' in regiment: unit = 'New Brunswick Provincial Regiment'
#    elif ' New Brunswick Fencibles ' in regiment: unit = 'New Brunswick Fencibles'
#    elif ' Nova Scotia Fencibles ' in regiment: unit = 'Nova Scotia Fencibles'
#    elif ' Royal Newfoundland Fencibles ' in regiment: unit = 'Royal Newfoundland Fencibles'
#    elif ' Royal West Indian Rangers' in regiment: unit = 'Royal West Indian Rangers'

    elif ' irish brigade' in regiment: unit = 'Irish Brigade'

#    elif ' Royal African Corps ' in regiment: unit = 'Royal African Corps'
#    elif ' Bourbon Regiment ' in regiment: unit = 'Bourbon Regiment'
#    elif ' Cape Regiment ' in regiment: unit = 'Cape Regiment'
#    elif ' Chasseurs Britanniques ' in regiment: unit = 'Chasseurs Britanniques'
#    elif ' Anglo-Corsican Regiment ' in regiment: unit = 'Anglo-Corsican Regiment'
#    elif ' Corsican Regiment ' in regiment: unit = 'Corsican Regiment'
#    elif " Dillon's Regiment " in regiment: unit = "Dillon's Regiment"
#    elif " Meuron's Regiment " in regiment: unit = "Meuron's Regiment"
#    elif " Roll's Regiment " in regiment: unit = "Roll's Regiment"
#    elif ' Royal Corsican Rangers ' in regiment: unit = 'Royal Corsican Rangers'
#    elif ' Royal Regiment of Malta ' in regiment: unit = 'Royal Regiment of Malta'
#    elif ' Sicilian Regiment ' in regiment: unit = 'Sicilian Regiment'
#    elif " Watteville's Regiment " in regiment: unit = "Watteville's Regiment"
#    elif ' York Chasseurs ' in regiment: unit = 'York Chasseurs'
#    elif ' York Light Infantry Volunteers ' in regiment: unit = 'York Light Infantry Volunteers'
#    elif ' York Rangers ' in regiment: unit = 'York Rangers'
#    elif ' 1st Regiment of Fencible Cavalry ' in regiment: unit = '1st Regiment of Fencible Cavalry'
#    elif ' 2nd Regiment of Fencible Cavalry ' in regiment: unit = '2nd Regiment of Fencible Cavalry'
#    elif ' Ancient British Fencible Cavalry ' in regiment: unit = 'Ancient British Fencible Cavalry'
#    elif ' Ayr Fencible Cavalry ' in regiment: unit = 'Ayr Fencible Cavalry'
#    elif ' Berskhire Fencible Cavalry ' in regiment: unit = 'Berskhire Fencible Cavalry'
#    elif ' Berwickshire Fencible Cavalry ' in regiment: unit = 'Berwickshire Fencible Cavalry'
#    elif ' Cambridgeshire Fencible Cavalry ' in regiment: unit = 'Cambridgeshire Fencible Cavalry'
#    elif ' Cinque Ports Fencible Cavalry ' in regiment: unit = 'Cinque Ports Fencible Cavalry'
#    elif ' Cornwall Fencible Cavalry ' in regiment: unit = 'Cornwall Fencible Cavalry'
#    elif ' Dumfriesshire Fencible Cavalry ' in regiment: unit = 'Dumfriesshire Fencible Cavalry'
#    elif ' Dunbarton Fencible Cavalry ' in regiment: unit = 'Dunbarton Fencible Cavalry'
#    elif ' Durham Fencible Cavalry ' in regiment: unit = 'Durham Fencible Cavalry'
#    elif ' East and West Lothian Fencible Cavalry ' in regiment: unit = 'East and West Lothian Fencible Cavalry'
#    elif ' Fifeshire Fencible Cavalry ' in regiment: unit = 'Fifeshire Fencible Cavalry'
#    elif ' Forfar Fencible Cavalry ' in regiment: unit = 'Forfar Fencible Cavalry'
#    elif ' Hampshire Fencible Cavalry ' in regiment: unit = 'Hampshire Fencible Cavalry'
#    elif ' Lanarkshire Fencible Cavalry ' in regiment: unit = 'Lanarkshire Fencible Cavalry'
#    elif ' Lancashire Fencible Cavalry ' in regiment: unit = 'Lancashire Fencible Cavalry'
#    elif ' Linlithgow Fencible Cavalry ' in regiment: unit = 'Linlithgow Fencible Cavalry'
#    elif ' Loyal Essex Fencible Cavalry ' in regiment: unit = 'Loyal Essex Fencible Cavalry'
#    elif ' Mid-Lothian Fencible Cavalry ' in regiment: unit = 'Mid-Lothian Fencible Cavalry'
#    elif ' New Romney Fencible Cavalry ' in regiment: unit = 'New Romney Fencible Cavalry'
#    elif ' Norfolk Fencible Cavalry ' in regiment: unit = 'Norfolk Fencible Cavalry'
#    elif ' Oxfordshire Fencible Cavalry ' in regiment: unit = 'Oxfordshire Fencible Cavalry'
#    elif ' Pembroke Fencible Cavalry ' in regiment: unit = 'Pembroke Fencible Cavalry'
#    elif ' Perthshire Fencible Cavalry ' in regiment: unit = 'Perthshire Fencible Cavalry'
#    elif " Princess Royal's Own Fencible Cavalry " in regiment: unit = "Princess Royal's Own Fencible Cavalry"
#    elif ' Roxburgh Fencible Cavalry ' in regiment: unit = 'Roxburgh Fencible Cavalry'
#    elif ' Rutland Fencible Cavalry ' in regiment: unit = 'Rutland Fencible Cavalry'
#    elif ' Somersetshire Fencible Cavalry ' in regiment: unit = 'Somersetshire Fencible Cavalry'
#    elif ' Surrey Fencible Cavalry ' in regiment: unit = 'Surrey Fencible Cavalry'
#    elif ' Sussex Fencible Cavalry ' in regiment: unit = 'Sussex Fencible Cavalry'
#    elif ' Warwickshire Fencible Cavalry ' in regiment: unit = 'Warwickshire Fencible Cavalry'
#    elif ' Strathspey Fencibles ' in regiment: unit = 'Strathspey Fencibles'
#    elif ' Sutherland Fencibles ' in regiment: unit = 'Sutherland Fencibles'
#    elif ' West Lowland Fencibles ' in regiment: unit = 'West Lowland Fencibles'
#    elif ' Perthshire Fencibles ' in regiment: unit = 'Perthshire Fencibles'
#    elif ' Argyllshire Fencibles ' in regiment: unit = 'Argyllshire Fencibles'
#    elif ' Northern Fencibles ' in regiment: unit = 'Northern Fencibles'
#    elif ' Southern Fencibles ' in regiment: unit = 'Southern Fencibles'
#    elif ' Rothsay and Caithness Fencibles ' in regiment: unit = 'Rothsay and Caithness Fencibles'
#    elif ' Ancient Irish Fencibles ' in regiment: unit = 'Ancient Irish Fencibles'
#    elif ' Angus Fencibles ' in regiment: unit = 'Angus Fencibles'
#    elif ' Angusshire Fencibles ' in regiment: unit = 'Angusshire Fencibles'
#    elif ' Caithness Legion ' in regiment: unit = 'Caithness Legion'
#    elif ' Cambrian Rangers ' in regiment: unit = 'Cambrian Rangers'
#    elif ' Cheshire Fencibles ' in regiment: unit = 'Cheshire Fencibles'
#    elif ' Devonshire and Cornwall Fencibles ' in regiment: unit = 'Devonshire and Cornwall Fencibles'
#    elif ' Banffshire Fencibles ' in regiment: unit = 'Banffshire Fencibles'
#    elif ' Dunbartonshire Fencibles ' in regiment: unit = 'Dunbartonshire Fencibles'
#    elif " Elgin's Fencibles " in regiment: unit = "Elgin's Fencibles"
#    elif ' Fifeshire Fencibles ' in regiment: unit = 'Fifeshire Fencibles'
#    elif ' Glengarry Fencibles ' in regiment: unit = 'Glengarry Fencibles'
#    elif ' Lochaber Fencibles ' in regiment: unit = 'Lochaber Fencibles'
#    elif ' Loyal British Fencibles ' in regiment: unit = 'Loyal British Fencibles'
#    elif ' Durham Fencibles ' in regiment: unit = 'Durham Fencibles'
#    elif ' Loyal Essex Fencibles ' in regiment: unit = 'Loyal Essex Fencibles'
#    elif ' Loyal Inverness Fencibles ' in regiment: unit = 'Loyal Inverness Fencibles'
#    elif ' Loyal Irish Fencibles ' in regiment: unit = 'Loyal Irish Fencibles'
#    elif ' Loyal Limerick Fencibles ' in regiment: unit = 'Loyal Limerick Fencibles'
#    elif ' Loyal Nottingham Fencibles ' in regiment: unit = 'Loyal Nottingham Fencibles'
#    elif ' Loyal Somerset Fencibles ' in regiment: unit = 'Loyal Somerset Fencibles'
#    elif ' Loyal Tarbert Fencbles ' in regiment: unit = 'Loyal Tarbert Fencbles'
#    elif ' Loyal Tay Fencibles ' in regiment: unit = 'Loyal Tay Fencibles'
#    elif ' North Lowland Fencibles ' in regiment: unit = 'North Lowland Fencibles'
#    elif ' Northampton Fencibles ' in regiment: unit = 'Northampton Fencibles'
#    elif ' Northumberland Fencibles ' in regiment: unit = 'Northumberland Fencibles'
#    elif ' Orkney and Shetland Fencibles ' in regiment: unit = 'Orkney and Shetland Fencibles'
#    elif ' Perth Fencibles ' in regiment: unit = 'Perth Fencibles'
#    elif " Prince of Wales's Fencibles " in regiment: unit = "Prince of Wales's Fencibles"
#    elif " Prince of Wales's Own Fencibles " in regiment: unit = "Prince of Wales's Own Fencibles"
#    elif " Princess of Wales's Fencibles " in regiment: unit = "Princess of Wales's Fencibles"
#    elif ' Reay Fencibles ' in regiment: unit = 'Reay Fencibles'
#    elif ' Ross and Cromarty Rangers ' in regiment: unit = 'Ross and Cromarty Rangers'
#    elif ' Ross-shire Fencibles ' in regiment: unit = 'Ross-shire Fencibles'
#    elif ' Royal Birmingham Fencibles ' in regiment: unit = 'Royal Birmingham Fencibles'
#    elif ' Royal Clan Alpine Fencibles ' in regiment: unit = 'Royal Clan Alpine Fencibles'
#    elif ' Royal Lancashire Volunteers ' in regiment: unit = 'Royal Lancashire Volunteers'
#    elif ' Royal Manx Fencibles ' in regiment: unit = 'Royal Manx Fencibles'
#    elif ' Scilly Isles Fencibles ' in regiment: unit = 'Scilly Isles Fencibles'
#    elif ' Shetland Fencibles ' in regiment: unit = 'Shetland Fencibles'
#    elif ' Suffolk Fencibles ' in regiment: unit = 'Suffolk Fencibles'
#    elif ' Surrey Rangers ' in regiment: unit = 'Surrey Rangers'
#    elif ' Loyal Macleod Fencibles ' in regiment: unit = 'Loyal Macleod Fencibles'
#    elif ' Regiment of the Isles ' in regiment: unit = 'Regiment of the Isles'
#    elif " Wallace's Fencibles " in regiment: unit = "Wallace's Fencibles"
#    elif ' York Fencibles ' in regiment: unit = 'York Fencibles'
#    elif ' 1st Provisional Battalion of Militia ' in regiment: unit = '1st Provisional Battalion of Militia'
#    elif ' 2nd Provisional Battalion of Militia ' in regiment: unit = '2nd Provisional Battalion of Militia'
#    elif ' 3rd Provisional Battalion of Militia ' in regiment: unit = '3rd Provisional Battalion of Militia'

    elif ' royal art' in regiment: unit = 'Royal Artillery'; number = ''
    elif ' r art' in regiment: unit = 'Royal Artillery'; number = ''

#    elif ' Royal Irish Artillery ' in regiment: unit = 'Royal Irish Artillery'
#    elif ' Royal Engineers ' in regiment: unit = 'Royal Engineers'
#    elif ' Royal Waggon Train ' in regiment: unit = 'Royal Waggon Train'
#    elif ' Royal Sappers and Miners ' in regiment: unit = 'Royal Sappers and Miners'
#    elif ' Royal Marines ' in regiment: unit = 'Royal Marines'
#    elif ' Royal Staff Corps ' in regiment: unit = 'Royal Staff Corps'

    elif ' light dragoon guards ' in regiment: unit = 'Light Dragoons'

    elif ' dragoon guards ' in regiment: unit = 'Dragoon Guards'
    elif ' drag guards ' in regiment: unit = 'Dragoon Guards'

    elif ' light dragoons ' in regiment: unit = 'Light Dragoons'

    elif ' dragoons ' in regiment: unit = 'Dragoons'

    elif ' foot guards ' in regiment: unit = 'Foot Guards'
    elif ' food guards ' in regiment: unit = 'Foot Guards'

    elif ' foot ' in regiment: unit = 'Foot'
    elif ' food ' in regiment: unit = 'Foot'

    elif ' dr ' in regiment: unit = 'Dragoons'

    elif ' f g ' in regiment: unit = 'Foot Guards'

    elif ' f ' in regiment: unit = 'Foot'

    #elif '' in regiment: unit = ''
    #elif '' in regiment: unit = ''
    #elif '' in regiment: unit = ''

    if unit == 'Royal Artillery' and ('irish' in regiment or 'ireland' in regiment): unit = 'Royal Irish Artillery'  

    regiment = regiment.replace('2d ', '2nd ')
    regiment = regiment.replace('3d ', '3rd ')

    if number == None:
        number = reg_number(regiment, number)

    if unit != None and number != None:

        regiment = number + unit  

    else:
    
        regiment = regiment.title()

        regiment = regiment.replace('1St ', '1st ')
        regiment = regiment.replace('2Nd ', '2nd ')
        regiment = regiment.replace('3Rd ', '3rd ')
        regiment = regiment.replace('4Th ', '4th ')
        regiment = regiment.replace('5Th ', '5th ')
        regiment = regiment.replace('6Th ', '6th ')
        regiment = regiment.replace('7Th ', '7th ')
        regiment = regiment.replace('8Th ', '8th ')
        regiment = regiment.replace('9Th ', '9th ')

        #regiment = orig_regiment

    regiment = regiment.strip()

    if regiment in unit_lookups:
        regiment = unit_lookups[regiment]


    if year != None:

        year = int(year)

        if regiment == '25th Light Dragoons' and year >= 1795 and year <= 1802: regiment = '22nd (formerly 25th) Light Dragoons'
        if regiment == '22nd Light Dragoons' and year >= 1803 and year <= 1820: regiment = '22nd (formerly 25th) Light Dragoons'
        if regiment == '26th Light Dragoons' and year >= 1795 and year <= 1802: regiment = '23rd (formerly 26th) Light Dragoons'
        if regiment == '23rd Light Dragoons' and year >= 1803 and year <= 1820: regiment = '23rd (formerly 26th) Light Dragoons'
        if regiment == '27th Light Dragoons' and year >= 1795 and year <= 1804: regiment = '24th (formerly 27th) Light Dragoons'
        if regiment == '24th Light Dragoons' and year >= 1805 and year <= 1820: regiment = '24th (formerly 27th) Light Dragoons'
        if regiment == '29th Light Dragoons' and year >= 1795 and year <= 1804: regiment = '25th (formerly 29th) Light Dragoons'
        if regiment == '25th Light Dragoons' and year >= 1805 and year <= 1820: regiment = '25th (formerly 29th) Light Dragoons'
        if regiment == '98th Foot' and year >= 1795 and year <= 1798: regiment = '91st (formerly 98th) Foot'
        if regiment == '91st Foot' and year >= 1799 and year <= 1820: regiment = '91st (formerly 98th) Foot'
        if regiment == '100th Foot' and year >= 1795 and year <= 1798: regiment = '92nd (formerly 100th) Foot'
        if regiment == '92nd Foot' and year >= 1799 and year <= 1820: regiment = '92nd (formerly 100th) Foot'
        if regiment == 'Scotch Brigade' and year >= 1793 and year <= 1803: regiment = '94th Foot (Scotch Brigade)'
        if regiment == '94th Foot' and year >= 1804 and year <= 1820: regiment = '94th Foot (Scotch Brigade)'
        if regiment == 'A Corps Of Riflemen' and year >= 1799 and year <= 1803: regiment = '95th Foot (The Rifles)'
        if regiment == '95th Foot' and year >= 1804 and year <= 1816: regiment = '95th Foot (The Rifles)'
        if regiment == 'The Rifle Brigade' and year >= 1817 and year <= 1820: regiment = '95th Foot (The Rifles)'
        if regiment == '1st Dragoons, KGL' and year >= 1804 and year <= 1813: regiment = '1st Heavy (later light) Dragoons, KGL'
        if regiment == '1st Light Dragoons, KGL' and year >= 1814 and year <= 1820: regiment = '1st Heavy (later light) Dragoons, KGL'
        if regiment == '2nd Dragoons, KGL' and year >= 1804 and year <= 1813: regiment = '2nd Heavy (later light) Dragoons, KGL'
        if regiment == '2nd Light Dragoons, KGL' and year >= 1814 and year <= 1820: regiment = '2nd Heavy (later light) Dragoons, KGL'
        if regiment == '1st Light Dragoons, KGL' and year >= 1804 and year <= 1813: regiment = '1st Light Dragoons (later Hussars), KGL'
        if regiment == '1st Hussars, KGL' and year >= 1814 and year <= 1820: regiment = '1st Light Dragoons (later Hussars), KGL'
        if regiment == '2nd Light Dragoons, KGL' and year >= 1804 and year <= 1813: regiment = '2nd Light Dragoons (later Hussars), KGL'
        if regiment == '2nd Hussars, KGL' and year >= 1814 and year <= 1820: regiment = '2nd Light Dragoons (later Hussars), KGL'
        if regiment == '3rd Light Dragoons, KGL' and year >= 1804 and year <= 1813: regiment = '3rd Light Dragoons (later Hussars), KGL'
        if regiment == '3rd Hussars, KGL' and year >= 1814 and year <= 1820: regiment = '3rd Light Dragoons (later Hussars), KGL'

    return regiment

























#def reg_regiment(regiment):
#
#    if regiment == None: return None
#    if isinstance(regiment, int): regiment = str(regiment)
#
#    # typos and irregularities
#
#    if regiment == 'Eighteenth Regiment of (Lion) Dragoons': regiment = 'Eighteenth Regiment of (Light) Dragoons'
#    if regiment == 'Fifteenth (or the Yorksh, Fast Riding) Regt. Of Foot': regiment = 'Fifteenth (or the Yorkshire, East Riding) Regiment of Foot'
#    if regiment == 'Fifteenth (or the Yorksh. East Riding) Regiment of  Foot': regiment = 'Fifteenth (or the Yorkshire, East Riding) Regiment of Foot'
#    if regiment == 'Fifty-fourth (or the West-Norsolk) Regt. Of Foot.': regiment = 'Fifty-Fourth (or the West Norfolk) Regiment Of Foot'
#    if regiment == 'Coldstream Regiment of Foot-Guards': regiment = 'Coldstream Regiment of Foot Guards'
#    if regiment == "Fifteenth (or the King's) Regt. Of (Light) Dragoon": regiment = "Fifteenth (or the King's) Regiment of (Light) Dragoons"
#    if regiment == "Fifty ninth (or the 2nd Nottinghamsh) Regt. Of Foot": regiment = "Fifty-Ninth (or the Second Nottinghamshire) Regiment of Foot"
#    if regiment == "First Regiment of Food-Guard": regiment = "First Regiment of Foot Guards"
#    if regiment == "First(or the Royal) Regiment of Foot(1st Battalien)": regiment = "First (or the Royal) Regiment of Foot (First Battalien)"
#    if regiment == "Forty Nine (or the Hertfordsh) Regt. Of Foot": regiment = "Forty-Ninth (or the Hertfordshire) Regiment of Foot"
#    if regiment == "Forty-ninth (or the Hertfordsh) Regt. Of Foot.": regiment = "Forty-Ninth (or the Hertfordshire) Regiment of Foot"
#    if regiment == "Fourteenth (or the Bedforshire) Regt. Of Foot": regiment = "Fourteenth (or the Bedfordshire) Regiment of Foot"
#    if regiment == "Fourth (or Royal Irish) Regiment of Dragoon Guards ": regiment = "Fourth (or the Royal Irish) Regiment of Dragoon Guards"
#    if regiment == "4th (or Royal Irish) Regiment of Dragoon Guards": regiment = "Fourth (or the Royal Irish) Regiment of Dragoon Guards"
#    if regiment == "Fourth (or the Queens own) Regt. Of Dragoons": regiment = "Fourth (or the Queen's Own) Regiment of Dragoons"
#    if regiment == "19th (or the 1st Yorksh. N. Riding) Regiment of Foot": regiment = "Nineteenth (or the First Yorkshire, North Riding) Regiment of Foot"
#    if regiment == "19th (or the Ist Yorksh. N. Riding) Regt.of Foot": regiment = "Nineteenth (or the First Yorkshire, North Riding) Regiment of Foot"
#    if regiment == "Ninteenth Regiment of (Light) Dragoons": regiment = "Nineteenth Regiment of (Light) Dragoons"
#    if regiment == "Second (or Royal N. Brit.) Regt. Of Dragoons": regiment = "Second (or Royal North British) Regiment Of Dragoons"
#    if regiment == "Second (or the Queen's) Regt. Of Foot": regiment = "Second (or the Queen's Royal) Regiment of Foot"
#    if regiment == "Seventh (or Princes's Royal's) Regt. Of Drag. Guards": regiment = "Seventh (or Princess Royal's) Regiment of Dragoon Guards"
#    if regiment == "Seventh (or The Queens own) Regiment of (Light) Dragoons": regiment = "Seventh (or the Queen's Own) Regiment of (Light) Dragoons"
#    if regiment == "Seventh Regt. Of Foot ( or Royal Fuzileers)": regiment = "Seventh Regiment of Foot (or the Royal Fuzileers)"
#    if regiment == "Seventh Regiment of Foot, (or the Royal Fuzileers)": regiment = "Seventh Regiment of Foot (or the Royal Fuzileers)"
#    if regiment == "Seventieth (or the Surry) Regiment of Foot.": regiment = "Seventieth (or the Surrey) Regiment of Foot"
#    if regiment == "Seventieth (or the Surry) Regiment of Foot": regiment = "Seventieth (or the Surrey) Regiment of Foot"
#    if regiment == "Sixteenth (or the Queens) Regt. Of (Light) Dragoons": regiment = "Sixteenth (or the Queen's) Regiment of (Light) Dragoons"
#    if regiment == "Sixty-eight (or the Durham) Regt. Of. Foot.": regiment = "Sixty-Eighth (or the Durham) Regiment of Foot"
#    if regiment == "65th (or the 2nd Yorkshire N. Riding) Regt. Of Foot": regiment = "Sixty-Fifth (or the Second Yorkshire, North Riding) Regiment of Foot"
#    if regiment == "65th (or the 2d Yorkshire n. Riding) Regt. Of Foot.": regiment = "Sixty-Fifth (or the Second Yorkshire, North Riding) Regiment of Foot"
#    if regiment == "61st (or the South Gloucestersh.) Regt. Of. Foot.": regiment = "Sixty-First (or the South Gloucestershire) Regiment of Foot"
#    if regiment == "Sixty Third (or the West Sussolk ) Regt. Of Foot": regiment = "Sixty-Third (or the West Suffolk) Regiment of Foot"
#    if regiment == "Tenth (or the Prince of Wales own) Regt. Of (Light) Dragoons": regiment = "Tenth (or the Prince of Wales's Own) Regiment of (Light) Dragoons"
#    if regiment == "Third (or Pr. Of Wales's) Regt. Of Dragoon Guards": regiment = "Third (or Prince of Wales's) Regiment of Dragoon Guards"
#    if regiment == "Third (or Pr.of Wales's) Regt. Of Dragoon Guarda": regiment = "Third (or Prince of Wales's) Regiment of Dragoon Guards"
#    if regiment == "Third (or the East King) Regiment of Foot,or the Buffs.": regiment = "Third (or the East Kent) Regiment of Foot, or the Buffs"
#    if regiment == "Third (or the Kings own) Regt. Of Dragoons": regiment = "Third (or the King's Own) Regiment of Dragoons"
#    if regiment == "Third Regiment of Foot-Guards": regiment = "Third Regiment of Foot Guards"
#    if regiment == "Thirteenth (or Ist Somersetshire) Regt. Of Foot": regiment = "Thirteenth (or the First Somersetshire) Regiment of Foot"
#    if regiment == "Thirteenth (or the 1st Smoersetshire) Regiment of Foot": regiment = "Thirteenth (or the First Somersetshire) Regiment of Foot"
#    if regiment == "38th ( or the 1st Staffordshire) Regt. Of Foot.": regiment = "Thirty-Eighth (or the First Staffordshire) Regiment of Foot"
#    if regiment == "Thirty Fifith (or the Doresetshire) Regt. Of Foot": regiment = "Thirty-Fifth (or the Dorsetshire) Regiment of Foot"
#    if regiment == "33rd (or the Ist Yorksh. West Riding ) Regt. Of Foot": regiment = "Thirty-Third (or the First Yorkshire, West Riding) Regiment of Foot"
#    if regiment == "33d (or the 1st Yorksh. West Riding) Regt. Of Foot": regiment = "Thirty-Third (or the First Yorkshire, West Riding) Regiment of Foot"
#    if regiment == "Twentieth (or the East Devonsh) Regt. Of Foot": regiment = "Twentieth (or the East Devonshire) Regiment of Foot"
#    if regiment == "28th (or the North Gloucestersh) Regt. Of Foot": regiment = "Twenty-Eighth (or the North Gloucestershire) Regiment of Foot"
#    if regiment == "28th  (or the North Gloucestersh) Regt. Of Foot": regiment = "Twenty-Eighth (or the North Gloucestershire) Regiment of Foot"
#    if regiment == "Twenty Feventh (or Inniskilling) Regt. Of Foot": regiment = "Twenty-Seventh (or Inniskilling) Regiment of Foot"
#    if regiment == "21st Regiment of Foot (or Royal N. Brit. Fuzileers)": regiment = "Twenty-First Regiment of Foot (or Royal North British Fuzileers)"
#    if regiment == "21st Regt. Of Foot (or Royal N. Brit. Fuzileers.)": regiment = "Twenty-First Regiment of Foot (or Royal North British Fuzileers)"
#    if regiment == "Twenty Fourth (or the 2nd Warwicksh) Regt. Of Foot": regiment = "Twenty-Fourth (or the Second Warwickshire) Regiment of Foot"
#    if regiment == "Eighteenth Regiment of (Light) Deagoons": regiment = "Eighteenth Regiment of (Light) Dragoons"
#    if regiment == "Eighth (or The Kings) Regiment of foot": regiment = "Eighth (or The King's) Regiment of Foot"
#    if regiment == "Second Regiment of Life Gurds": regiment = "Second Regiment of Life Guards"
#    if regiment == "Fifth (or the Northumberland) Regt.of foot": regiment = "Fifth (or the Northumberland) Regiment of Foot"
#    if regiment == "Fifth Regiment of Dragon-Guards": regiment = "Fifth Regiment of Dragoon Guards"
#    if regiment == "Fifty-fifth (or the Westmoreland ) Regt . Of Foot": regiment = "Fifty-Fifth (or the Westmoreland) Regiment of Foot"
#    if regiment == "First (or the King's) Regt.of Dragoon-Guards": regiment = "First (or the King's) Regiment of Dragoon Guards"
#    if regiment == "First (or Royal) Regiment of Dragooons": regiment = "First (or the Royal) Regiment of Dragoons"
#    if regiment == "First (or the Royal) Regt of Foot (1st Bat)": regiment = "First (or the Royal) Regiment of Foot (First Battalien)"
#    if regiment == "First (or the Royal) Regt of Foot (2nd Bat)": regiment = "First (or the Royal) Regiment of Foot (Second Battalien)"
#    if regiment == "First Regiment of Life Gurds": regiment = "First Regiment of Life Guards"
#    if regiment == "Fourth (or the Queen's own) Regt.of Dragoons.": regiment = "Fourth (or the Queen's Own) Regiment of Dragoons"
#    if regiment == "19th (or the 1st Yorksh. N. Riding) Regt. of Foot ": regiment = "Nineteenth (or the First Yorkshire, North Riding) Regiment of Foot"
#    if regiment == "Royal Regiment of Horse-Guards": regiment = "Royal Regiment of Horse Guards"
#    if regiment == "Second (or Royal N. Brit.) Regt. of Dragooons": regiment = "Second (or Royal North British) Regiment of Dragoons"
#    if regiment == "Second (or the Queen's)  Regt. Of Dragoon-Guards": regiment = "Second (or the Queen's) Regiment of Dragoon Guards"
#    if regiment == "Second (or the Queens Royal) Regt Of Foot": regiment = "Second (or the Queen's Royal) Regiment of Foot"
#    if regiment == "Seventh (or Princess Royal's) Regt.of Drag. Guards.": regiment = "Seventh (or Princess Royal's) Regiment of Dragoon Guards"
#    if regiment == "Seventh Regt. of Foot (or Royal Fuzileers.)": regiment = "Seventh Regiment of Foot (or the Royal Fuzileers)"
#    if regiment == "Seventieth (or the Surry) Regt. Of Foot": regiment = "Seventieth (or the Surrey) Regiment of Foot"
#    if regiment == "Sixteenth (or the Buckingham) Regt of Foot": regiment = "Sixteenth (or the Buckinghamshire) Regiment of Foot"
#    if regiment == "Sixth Regiment of Dragon-Guards": regiment = "Sixth Regiment of Dragoon Guards"
#    if regiment == "61st (or the South Gloucestersh) Regt. Of Foot": regiment = "Sixty-First (or the South Gloucestershire) Regiment of Foot"
#    if regiment == "Sixty-fourth (or the 2d Staffordsh) Regt. Of Foot.": regiment = "Sixty-Fourth (or the Second Staffordshire) Regiment of Foot"
#    if regiment == "Sixty Fourth (or the 2nd Staffordsh) Regt. Of Foot": regiment = "Sixty-Fourth (or the Second Staffordshire) Regiment of Foot"
#    if regiment == "Tenth (or the Prince of Wales Own) Regt of (Light) Dragoons": regiment = "Tenth (or the Prince of Wales's Own) Regiment of (Light) Dragoons"
#    if regiment == "Third (or Pr. Of Wales's)  Regt. Of Dragoon-Guards": regiment = "Third (or Prince of Wales's) Regiment of Dragoon Guards"
#    if regiment == "Third (or the East King) Regt. Of Foot, or the Buffs": regiment = "Third (or the East Kent) Regiment of Foot, or the Buffs"
#    if regiment == "Third (or the King's own) Regt.of Dragoons.": regiment = "Third (or the King's Own) Regiment of Dragoons"
#    if regiment == "Third Regiment of Foot-Guards.": regiment = "Third Regiment of Foot Guards"
#    if regiment == "33rd (or the 1st Yorksh. West Riding) Regt. of Foot": regiment = "Thirty-Third (or the First Yorkshire, West Riding) Regiment of Foot"
#    if regiment == "Twelth (or the Prince of Wales's) Regt of (Light) Dragoons": regiment = "Twelfth (or Prince of Wales's) Regiment of (Light) Dragoons"
#    if regiment == "Twentieth (or the East Devonsh) Regt. of Foot": regiment = "Twentieth (or the East Devonshire) Regiment of Foot"
#    if regiment == "Twenty-fifth (or the Suffolk) Regiment of Foot": regiment = "Twenty-Fifth (or the Sussex) Regiment of Foot"
#    if regiment == "21st Regt. of Foot (or Royal N. Brit. Fuzileers)": regiment = "Twenty-First Regiment of Foot (or Royal North British Fuzileers)"
#    if regiment == "Twenty-seventh ( or Inniskilling) Regt of Foot": regiment = "Twenty-Seventh (or Inniskilling) Regiment of Foot"
#    if regiment == "23d Regt of (or Royal Welsh Fuzileers)": regiment = "Twenty-Third Regiment of Foot (or Royal Welsh Fuzileers)"
#    if regiment == "Eighteen Regiment of (Light) Dragoons": regiment = "Eighteenth Regiment of (Light) Dragoons"
#    if regiment == "Fifteenth (or the Yorksh East Riding) Regt of Foot": regiment = "Fifteenth (or the Yorkshire, East Riding) Regiment of Foot"
#    if regiment == "Third (or the East Kent) Regiment of Foot,or the Buffs.": regiment = "Third (or the East Kent) Regiment of Foot, or the Buffs"
#    if regiment == "Twenty Fifth  (or the Suffolk) Regt. Of Foot": regiment = "Twenty-Fifth (or the Sussex) Regiment of Foot"
#
#    regiment = regiment.replace(' ) ', ') ')
#
#    regiment = regiment.replace('Seventy- ', 'Seventy-')
#
#    regiment = regiment.replace(' (or Royal) ', ' (or the Royal) ')
#
#    regiment = re.sub(r'\bRegt\.', 'Regiment', regiment)
#
#    regiment = re.sub(r'\bRegt\b', 'Regiment', regiment)
#
#    regiment = re.sub(r'\bBat\.', 'Battalien', regiment)
#
#    regiment = re.sub(r'\bDrag\.', 'Dragoon', regiment)
#
#    regiment = re.sub(r'\bYorksh\.', 'Yorkshire', regiment)
#
#    regiment = re.sub(r'\bBuckinghamsh\.', 'Buckinghamshire', regiment)
#
#    regiment = re.sub(r'\bNottinghamsh\.', 'Nottinghamshire', regiment)
#
#    regiment = re.sub(r'\bGloucestersh\.', 'Gloucestershire', regiment)
#
#    regiment = re.sub(r'\bDevonsh\.', 'Devonshire', regiment)
#
#    regiment = re.sub(r'\bWarwicksh\.', 'Warwickshire', regiment)
#
#    regiment = re.sub(r'\bHertfordsh\.', 'Hertfordshire', regiment) 
#
#    regiment = re.sub(r'\bStaffordsh\.', 'Staffordshire', regiment) 
#
#    regiment = re.sub(r'\bFifith\b', 'Fifth', regiment)
#
#    regiment = re.sub(r'\b[1I]st\b', 'First', regiment)
#
#    regiment = re.sub(r'\b2n?d\b', 'Second', regiment)
#
#    regiment = re.sub(r'\b23r?d\b', 'Twenty-Third', regiment)
#
#    regiment = re.sub(r'\b33r?d\b', 'Thirty-Third', regiment)
#
#    regiment = re.sub(r'\b37[nrst][dth]\b', 'Thirty-Seventh', regiment)
#
#    regiment = re.sub(r'\b42n?d\b', 'Forty-Second', regiment)
#
#    regiment = re.sub(r'\b43r?d\b', 'Forty-Third', regiment)
#    regiment = re.sub(r'\b4th\b', 'Fourth', regiment)
#    regiment = re.sub(r'\b19th\b', 'Nineteenth', regiment)
#    regiment = re.sub(r'\b21st\b', 'Twenty-First', regiment)                                                   
#    regiment = re.sub(r'\b28th\b', 'Twenty-Eighth', regiment)                                                      
#    regiment = re.sub(r'\b38th\b', 'Thirty-Eighth', regiment)                                                       
#    regiment = re.sub(r'\b39th\b', 'Thirty-Ninth', regiment)                                                         
#    regiment = re.sub(r'\b40th\b', 'Fortieth', regiment)                                                           
#    regiment = re.sub(r'\b44th\b', 'Forty-Fourth', regiment)                                                           
#    regiment = re.sub(r'\b45th\b', 'Forty-Fifth', regiment)                                                        
#    regiment = re.sub(r'\b46th\b', 'Forty-Sixth', regiment)                                                        
#    regiment = re.sub(r'\b47th\b', 'Forty-Seventh', regiment)                                                        
#    regiment = re.sub(r'\b48th\b', 'Forty-Eighth', regiment)                                                       
#    regiment = re.sub(r'\b51st\b', 'Fifty-First', regiment)                                              
#    regiment = re.sub(r'\b51th\b', 'Fifty-First', regiment)                                              
#    regiment = re.sub(r'\b57th\b', 'Fifty-Seventh', regiment)                                                          
#    regiment = re.sub(r'\b61st\b', 'Sixty-First', regiment)                                                   
#    regiment = re.sub(r'\b65th\b', 'Sixty-Fifth', regiment)                                                 
#    regiment = re.sub(r'\b67th\b', 'Sixty-Seventh', regiment)                                                        
#    regiment = re.sub(r'\b69th\b', 'Sixty-Ninth', regiment)
#
#    regiment = re.sub(r'\bTwenty ([FSTENfsten])', r'Twenty-\1', regiment)
#    regiment = re.sub(r'\bThirty ([FSTENfsten])', r'Thirty-\1', regiment)
#    regiment = re.sub(r'\bForty ([FSTENfsten])', r'Forty-\1', regiment)
#    regiment = re.sub(r'\bFifty ([FSTENfsten])', r'Fifty-\1', regiment)
#    regiment = re.sub(r'\bSixty ([FSTENfsten])', r'Sixty-\1', regiment)
#    regiment = re.sub(r'\bSeventy ([FSTENfsten])', r'Seventy-\1', regiment)
#
#    regiment = re.sub(r' +', ' ', regiment)
#
#    regiment = regiment.title()
#
#    regiment = re.sub(r"'S\b", "'s", regiment)
#
#    regiment = re.sub(r'\bOf\b', 'of', regiment)
#
#    regiment = re.sub(r'\bOr\b', 'or', regiment)
#
#    regiment = re.sub(r'\bThe\b', 'the', regiment)
#
#    regiment = regiment.strip()
#
#    regiment = regiment.rstrip('.')  
#
#    if (len(regiment) > 0): regiment = regiment[0].upper() + regiment[1:]
#
#    return regiment



def reg_rank(rank):

    if rank == None: return None

    if isinstance(rank, int): rank = str(rank)

    # typos and irregularities
    if rank.strip() in rank_lookups_unregularised: rank = rank_lookups_unregularised[rank.strip()]

    if 'Quarter - Masster' in rank: rank = 'Quarter-Master'
    if 'Quarter Mafter' in rank: rank = 'Quarter-Master'
    if 'Quarter Master' in rank: rank = 'Quarter-Master'
    if 'Quarter- Master' in rank: rank = 'Quarter-Master'
    if 'Quarter-Master' in rank: rank = 'Quarter-Master'
    if 'Quartermaster' in rank: rank = 'Quarter-Master'
    if 'Q Mr' in rank: rank = 'Quarter-Master'
    if 'Pay-Master' in rank: rank = 'Paymaster'

    if rank == 'Solictior': rank = 'Solicitor'
    if rank == 'Surgen': rank = 'Surgeon'
    if rank == 'Surgoen': rank = 'Surgeon' 
    if rank == 'Surgeon of a Recruiting District': rank = 'Surgeon'

    rank = rank.replace('1st ', 'First ')
    rank = rank.replace('1 st ', 'First ')
    rank = rank.replace('Frist ', 'First ')
    rank = rank.replace('2nd ', 'Second ')
    rank = rank.replace('2d ', 'Second ')

    rank = rank.replace('Adjut. ', 'Adjutant ')
    rank = rank.replace('Adjut ', 'Adjutant ')
    rank = rank.replace('Assist. ', 'Assistant ')
    rank = rank.replace('Affift ', 'Assistant ')

    rank = rank.replace('Lieutentant', 'Lieutenant')

    rank = rank.replace('Lieut. ', 'Lieutenant ')
    rank = rank.replace('Lieuts.', 'Lieutenant')
    rank = rank.replace('Lieuts', 'Lieutenant')
    rank = rank.replace('Lieut.', 'Lieutenant')
    rank = rank.replace(' Leiut. ', ' Lieutenant ')
    rank = re.sub(r'\bLieut\b', 'Lieutenant', rank)                                                       
    rank = rank.replace('Leut. ', 'Lieutenant ')
    rank = rank.replace('Lt.', 'Lieutenant ')
    rank = rank.replace(' Lt ', ' Lieutenant ')

    rank = rank.replace('Comm.', 'Commandant')

    rank = re.sub(r'\bComm\b', 'Commandant', rank)

    rank = rank.replace('Captains', 'Captain')
    rank = rank.replace('Captian', 'Captain')
    rank = rank.replace('Capatian', 'Captain')
    rank = rank.replace('Capatain', 'Captain')
    rank = rank.replace('capatain', 'Captain')
    rank = rank.replace('Captan', 'Captain')
    rank = rank.replace('Capt.', 'Captain')
    
    rank = re.sub(r'\bCapt\b', 'Captain', rank)

    rank = re.sub(r'\bAdv\b', 'Advocate', rank)
    rank = re.sub(r'\bDep\b', 'Deputy', rank)

    rank = rank.replace('Lieutatant', 'Lieutenant')
    rank = rank.replace('Lieutenent', 'Lieutenant')
    rank = rank.replace('lieutenant', 'Lieutenant')
    rank = rank.replace('Leutenant', 'Lieutenant')
    rank = rank.replace('Leiutenant', 'Lieutenant')
    rank = rank.replace('Lientenant', 'Lieutenant')
    rank = rank.replace('Lieuenant', 'Lieutenant')
    rank = rank.replace('Lieutanant', 'Lieutenant')
    rank = rank.replace('Lieutetant', 'Lieutenant')

    rank = rank.replace('Marshall', 'Marshal')

    rank = rank.replace('Col.', 'Colonel')
    rank = re.sub(r'\bCol\b', 'Colonel', rank)
    rank = rank.replace('Coloael', 'Colonel')
    rank = rank.replace('Cononel', 'Colonel')

    rank = rank.replace('Commissaries', 'Commissary')
    rank = rank.replace('COMMISSARIES', 'Commissary')
    rank = rank.replace('Commissarie', 'Commissary')

    rank = rank.replace('Surgeons', 'Surgeon')
    rank = rank.replace('Chaplains', 'Chaplain')
    rank = rank.replace('Inspectors', 'Inspector')

    rank = rank.replace('Deputy -', 'Deputy ')
    rank = rank.replace('Assistant-', 'Assistant ')

    rank = rank.replace('Lieutenant Colonel', 'Lieutenant-Colonel')
    rank = rank.replace('Lieutenant -Colonel', 'Lieutenant-Colonel')
    rank = rank.replace('Lieutenant- Colonel', 'Lieutenant-Colonel')

    rank = rank.replace('Lieutenant General', 'Lieutenant-General')

    rank = rank.replace('-Local Rank', '(Local Rank)')

    rank = rank.replace('Field Marshal', 'Field-Marshal')

    rank = rank.replace('major general', 'Major-General')
    rank = rank.replace('Major General', 'Major-General')
    #rank = rank.replace('Major- General', 'Major-General')

    rank = rank.replace('Gen.', 'General')
    rank = re.sub(r'\bGen\b', 'General', rank)

    rank = rank.replace('Maj.', 'Major')
    rank = re.sub(r'\bMaj\b', 'Major', rank)

    rank = rank.replace('Lieutenant Colonel', 'Lieutenant-Colonel')

    rank = rank.replace('Captain Commissary', 'Captain-Commissary')
    rank = rank.replace('Lieutenant Commissary', 'Lieutenant-Commissary')
    rank = rank.replace('Quartermaster', 'Quarter-master')
    rank = rank.replace('Brigadier General', 'Brigadier-General')

    rank = rank.replace('Additonal ', 'Additional ')

    rank = rank.replace('.', ' ')

    rank = rank.replace('- ', '-')
    rank = rank.replace(' – ', '-')

    rank = rank.replace('&', ' and ')

    rank = re.sub(r' +', ' ', rank) # regularise space

    rank = rank.title()

    rank = rank.replace(' And ', ' and ')

    rank = rank.replace(' In ', ' in ')

    rank = rank.replace(' Of ', ' of ')

    # Some Captaincies say where they are capatin of eg 'Captain of Sandown'
    if rank.startswith('Captain Of '): rank = 'Captain'

    rank = rank.replace('Store-Keepers', 'Store-Keeper')
    rank = rank.replace('Storekeeper', 'Store-Keeper')
    rank = rank.replace('Store keeper', 'Store-Keeper')

    rank = rank.replace('Colonel-Commandant', 'Colonel Commandant')

    rank = rank.strip()

    rank = rank.rstrip('s') # No recognised ranks end in 's' and sheet 1 sometimes pluralises ranks, so we remove any trailing s

    rank = re.sub(r'\bProvision$', 'Provisions', rank)
    rank = re.sub(r'\bHospital$', 'Hospitals', rank)

    if rank.strip() in rank_lookups_regularised: rank = rank_lookups_regularised[rank.strip()]

    return rank



def reg_date(date):

    reg_date = None

    if date != None:

        if isinstance(date, datetime): return date.date()
        if isinstance(date, int): date = str(date)
        if isinstance(date, float): date = str(date)

        #print(date)

        # typos and irregularities

        if date == '13 December 783': date = '13 December 1783'
        if date == '31 August 1790s': date = '31 August 1790'
        if date == 'Gen. 20 November 1782': date = '20 November 1782'
        if date == '06 April 791': date = '6 April 1791'
        if date == '13 December !782': date = '13 December 1782'
        if date == '14 March !789': date = '14 March 1789'
        if date == '15b June 1791': date = '15 June 1791'
        if date == '24 October 11789': date = '24 October 1789'
        if date == '25 June !788': date = '25 June 1788'
        if date == '31 March 17888': date = '31 March 1788'
        if date == '4th July 1788': date = '4 July 1788'
        if date == '': date = ''
        if date == '': date = ''
        if date == '': date = ''
        if date == '': date = ''
        if date == '': date = ''
        if date == '': date = ''
        if date == '': date = ''
        if date == '': date = ''

        date = re.sub(r'[^0-9a-zA-Z ]', r'', date)

        date = re.sub(r'([0-9])([A-Z])', r'\1 \2', date)

        date = re.sub(r'([a-z])([0-9])', r'\1 \2', date)

        date = date.replace(' Janiuary ', ' January ')
        date = date.replace(' Janurary ', ' January ')
        date = date.replace(' Januaey ', ' January ')
        date = date.replace(' Jan ', ' January ')
        date = date.replace(' jan ', ' January ')

        date = date.replace(' Frebruary ', ' February ')
        date = date.replace(' Febrauary ', ' February ')
        date = date.replace(' Febryary ', ' February ')
        date = date.replace(' Febraury ', ' February ')
        date = date.replace(' February ', ' February ')
        date = date.replace(' Fenruary ', ' February ')
        date = date.replace(' Februar ', ' February ')
        date = date.replace(' Feb ', ' February ')

        date = date.replace(' Mar ', ' March ')

        date = date.replace('14april ', '14 April ')
        date = date.replace('14april ', '14 April ')
        date = date.replace(' Aprili ', ' April ')
        date = date.replace(' Apr ', ' April ')

        date = date.replace(' Mau ', ' May ')

        date = date.replace(' Junr ', ' June ')
        date = date.replace(' Jun ', ' June ')

        date = date.replace(' Jul ', ' July ')

        date = date.replace('14august ', '14 August ')
        date = date.replace(' Augustt ', ' August ')
        date = date.replace(' Auguts ', ' August ')
        date = date.replace(' Augsut ', ' August ')
        date = date.replace(' Aug ', ' August ')

        date = date.replace(' Septembermber ', ' September ')
        date = date.replace(' sepetember ', ' September ')
        date = date.replace(' Septembere ', ' September ')
        date = date.replace(' Sepetember ', ' September ')
        date = date.replace(' Spetember ', ' September ')
        date = date.replace(' Sepetmber ', ' September ')
        date = date.replace(' Spetmber ', ' September ')
        date = date.replace(' Sept ', ' September ')
        date = date.replace(' Sep ', ' September ')

        date = date.replace(' Octoberobter ', ' October ')
        date = date.replace(' Octpmber ', ' October ')
        date = date.replace(' Octomber ', ' October ')
        date = date.replace(' octomber ', ' October ')
        date = date.replace(' ocotober ', ' October ')
        date = date.replace(' ocotber ', ' October ')
        date = date.replace(' Octomer ', ' October ')
        date = date.replace(' Ocotber ', ' October ')
        date = date.replace(' Oct ', ' October ')

        date = date.replace(' Nobvember ', ' November ')
        date = date.replace(' Novvember ', ' November ')
        date = date.replace(' Novemver ', ' November ')
        date = date.replace(' November ', ' November ')
        date = date.replace(' Novenber ', ' November ')
        date = date.replace(' Novmeber ', ' November ')
        date = date.replace(' Nobember ', ' November ')
        date = date.replace(' Novembe ', ' November ')
        date = date.replace(' Novmber ', ' November ')
        date = date.replace(' Nov ', ' November ')

        date = date.replace(' Decemberembe ', ' December ')    
        date = date.replace(' Decemeber ', ' December ')
        date = date.replace(' Dcember ', ' December ')
        date = date.replace(' Decmber ', ' December ')    
        date = date.replace(' Dec ', ' December ')    

        date = date.strip()

        try:

            reg_date = datetime.strptime(date, '%d %B %Y').date()

        except Exception as ex:

            pass
            #print('Error parsing ' + str(date))
            #print(ex)
            #print(traceback.format_exc())
            #exit(1)
            
        #print(reg_date)

    return reg_date


def reg_name(name):

    if isinstance(name, int): name = str(name)

    name = name.lower()

    #name = name.replace("_", ". ") # Names in sheet 10 (Regiments) in 1815 and possibly other places are littered with underscores which seem to follow initials.
    #name = name.replace("-", " ") # Names in sheet 10 (Regiments) in 1815 and possibly other places are littered with dashes which seem to represent multiple things.
 
    name = re.sub(r'[-_\.]([^ ])', r' \1', name) # Kevin: "I think the logic would be that we replace – or _  (and probably a . too) with a space
                                                 # if there isn’t a space already after it; otherwise it can be deleted."
                                                 # See the two lines above for the old approach.
                                                 # The line below will remove the characters not caught by this rule.

    name = re.sub(r'[^a-z\., ]', r'', name) # Remove every character which is not whitelisted.

    name = re.sub(r' +', ' ', name) # Regularise space

    name = name.replace("' ", "'") # We want to avoid L' Estrange, O' Loghlin etc should be L'Estrange, O'Loghlin etc

    name = re.sub(r'\.([^ ])', r'. \1', name) # Enforce space after every full stop.
    name = re.sub(r',([^ ])', r', \1', name) # Enforce space after every comma. 

    name = name.replace(' .', '.') # Enforce no spaces before full stops.
    name = name.replace(' ,', ',') # Enforce no spaces before commas.


    # Capture and correct surname first names eg. 'Hastings, Gen. Sir Charles'
    surname_first_match = re.search('^([a-z]*), ', name)
    if (surname_first_match):
        #print('surname first match')
        surname = surname_first_match.group(1)
        #print('**' + surname + '**')
        name = name[len(surname) + 2:] + ' ' + surname
        #print(name)
        #exit()

    name = name.replace('wm. ', 'william ')
    name = name.replace('will. ', 'william ')
    name = name.replace('wiliam ', 'william ')
    name = name.replace('wiiliam ', 'william ')
    name = name.replace('wiliim ', 'william ')
    name = name.replace('will ', 'william ')
    name = name.replace('willaim ', 'william ')
    name = name.replace('willia ', 'william ')
    name = name.replace('william ', 'william ')
    name = name.replace('williams ', 'william ')
    name = name.replace('willliam ', 'william ')
    name = name.replace('tho. ', 'thomas ')
    name = name.replace('geo. ', 'george ')
    name = name.replace('cha. ', 'charles ')
    name = name.replace('ch. ', 'charles ')
    name = name.replace('rich. ', 'richard ')
    name = name.replace('thos. ', 'thomas ')
    name = name.replace('tho ', 'thomas ')
    name = name.replace('jn. ', 'john ')
    name = name.replace('fred. ', 'frederick ')
    name = name.replace('fitz. ', 'fitzwilliam ')
    name = name.replace('fitz ', 'fitzwilliam ')
    name = name.replace('fiz. william ', 'fitzwilliam ')
    name = name.replace('edw. ', 'edward ')
    name = name.replace('ed. ', 'edward ')
    name = name.replace('hen. ', 'henry ')
    name = name.replace('ben. ', 'benjamin ')
    name = name.replace('benj. ', 'benjamin ')
    name = name.replace('alex. ', 'alexander ')
    name = name.replace('rob. ', 'robert ')
    name = name.replace('rd. ', 'richard ')
    name = name.replace('jos. ', 'joseph ')

    name = name.replace('hon. ', 'honourable ')

    name = name.replace(' marq. ', ' marquis ')
    name = name.replace(' marq ', ' marquis ')
    name = name.replace(' mar. ', ' marquis ')
    name = name.replace(' m. of ', ' marquis of ')
    name = name.replace(' m of ', ' marquis of ')
    name = name.replace(' e. of ', ' earl of ')
    name = name.replace(' e of ', ' earl of ')
    name = name.replace(' d. of ', ' duke of ')
    name = name.replace(' d of ', ' duke of ')

    name = name.replace(' lord. ', ' lord ') # full stop after lord is not allowed
    name = name.replace('ld. ', 'lord ')
    name = re.sub(r'([a-z]) lord', r'\1, lord', name) # Comma must be between given name and Lord eg 'Francis Lord Heathfield' must become 'Francis, Lord Heathfield'

    name = name.replace('&', '') # Ampersands sit in honourifics and make them harder to remove

    name = re.sub(r'\b([a-z]) ', r'\1. ', name) # Single orphan characters eg t get a full stop after them eg t. (cannot use \b at end because full stop is itself a word terminator
    name = re.sub(r'\b([a-z])$', r'\1.', name) # Single orphan characters eg t get a full stop after them eg t. (at end of string)

    name = re.sub(r' +', ' ', name) # Regularise space again

    name = name.title()

    name = name.strip()

    # Some names start with ranks
    name_rank = ''
    if name.startswith('Maj. Gen. '): name = name[10:]; name_rank = 'Major-General'
    if name.startswith('Lt. Gen. '): name = name[9:]; name_rank = 'Lieutenant-General'
    if name.startswith('Lt. Col. '): name = name[9:]; name_rank = 'Lieutenant-Colonel'
    if name.startswith('Colonel '): name = name[8:]; name_rank = 'Colonel'
    if name.startswith('General '): name = name[8:]; name_rank = 'General'
    if name.startswith('M. Gen. '): name = name[8:]; name_rank = 'Major-General'
    if name.startswith('Captain '): name = name[8:]; name_rank = 'Captain'
    if name.startswith('Lieut. '): name = name[7:]; name_rank = 'Lieutenant'
    if name.startswith('Ensign '): name = name[7:]; name_rank = 'Ensign'
    if name.startswith('Capt. '): name = name[6:]; name_rank = 'Captain'
    if name.startswith('Col. '): name = name[5:]; name_rank = 'Colonel'
    if name.startswith('Gen. '): name = name[5:]; name_rank = 'General'
    if name.startswith('Gn. '): name = name[4:]; name_rank = 'General'
    if name.startswith('Lt. '): name = name[4:]; name_rank = 'Lieutenant'


    # HumanName needs the name to be processed more in order to work properly. phname = prepared for HumanName
    phname = name

    phname = phname.replace('His Serene Highness', '')
    phname = phname.replace('His Royal Highness', '')
    phname = phname.replace('His R. H.', '')
    phname = phname.replace('Royal Highness', '')
    phname = phname.replace('K. G. G. C. B.', '')
    phname = phname.replace('G. G. C. B.', '')
    phname = phname.replace('K. C. B.', '')
    phname = phname.replace('K. St. P.', '')
    phname = phname.replace('G. G. C. B.', '')
    phname = phname.replace('G. C. B.', '')
    phname = phname.replace('Gcb.', '')
    phname = phname.replace('Kst. P', '')
    phname = phname.replace('Kstp', '')
    phname = phname.replace('Bt.', '')
    phname = phname.replace('K. G.', '')
    phname = phname.replace('K. B.', '')
    phname = phname.replace('K. T.', '')
    phname = phname.replace(', M. D.', '')
    phname = phname.replace(', M.D.', '')
    phname = phname.replace(',M.D.', '')
    phname = phname.replace(',M. D.', '')
    phname = phname.replace(',', '') # Usually detects name position better with commas removed
    phname = phname.replace('Lord', '') # Usually detects name position better with lord removed
    phname = phname.strip()
    if phname.endswith('.'): phname = phname[:-1] # Usually detects name position better with any following full stops removed
    phname = phname.strip()

    #print('     phname: **' + phname + '**')

    hname = HumanName(phname)

    surname = hname.last
    given = hname.first
    middlenames = hname.middle
    title = hname.title
    namesuffix = hname.suffix
    nickname = hname.nickname

    if len(surname) == 0 and ' ' not in name: # HumanName interprets surname only names as given names (for example see sheet 34 (Casualties) 1815)
        surname = name
        given = ''

    return name, surname, given, middlenames, title, namesuffix, nickname, name_rank



date2_rank_abbreviations = {}

date2_rank_abbreviations['field marshal'] = 'Field-Marshal'
date2_rank_abbreviations['f mar'] = 'Field-Marshal'

date2_rank_abbreviations['gen'] = 'General'

date2_rank_abbreviations['lieutenant general'] = 'Lieutenant-General'
date2_rank_abbreviations['lt gen'] = 'Lieutenant-General'
date2_rank_abbreviations['l g'] = 'Lieutenant-General'

date2_rank_abbreviations['m gen'] = 'Major-General'
date2_rank_abbreviations['mgen'] = 'Major-General'
date2_rank_abbreviations['m g'] = 'Major-General'

date2_rank_abbreviations['col'] = 'Colonel'

date2_rank_abbreviations['lieutenant colonel'] = 'Lieutenant-Colonel'
date2_rank_abbreviations['lt co'] = 'Lieutenant-Colonel'
date2_rank_abbreviations['l co'] = 'Lieutenant-Colonel'

date2_rank_abbreviations['maj'] = 'Major'

date2_rank_abbreviations['cap lt'] = 'Captain Lieutenant'

date2_rank_abbreviations['cap'] = 'Captain'

date2_rank_abbreviations['lieutenant'] = 'Lieutenant'
date2_rank_abbreviations['lieut'] = 'Lieutenant'
date2_rank_abbreviations['lt'] = 'Lieutenant'

date2_rank_abbreviations['cornet'] = 'Cornet'

date2_rank_abbreviations['ensign'] = 'Ensign'

#date2_rank_abbreviations[''] = ''
#date2_rank_abbreviations[''] = ''
#date2_rank_abbreviations[''] = ''
#date2_rank_abbreviations[''] = ''

def reg_date2(date2):

    regd = None

    if isinstance(date2, str):

        
        date2 = date2.lower()
        date2 = re.sub(r'[^0-9a-z ]', r' ', date2) # Replace any character which is not a lower case letter, a number or a space, with a space.
        date2 = re.sub(r' +', ' ', date2) # Regularise space.

        for rakey in date2_rank_abbreviations:

            if date2.startswith(rakey):

                #print('Had term')

                #print(date2)

                date_string = date2[len(rakey):]
                date_string = date_string.strip()

                #print('**' + date2_rank_abbreviations[rakey] + '**')
                #print('**' + date_string + '**')
                regd = reg_date(date_string)
                #print(regd)
                date2rank = date2_rank_abbreviations[rakey]
                date2lvl = rank_lvl(date2rank)

                if date2lvl is None:

                    print('reg_date2 assigned an invalid rank: **' + date2rank + '**')
                    exit()

                return regd, date2rank, date2lvl



        # Maybe the date2 string is just a date with no rank?

        regd = reg_date(date2)

    return regd, None, None

