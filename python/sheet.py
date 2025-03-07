#!/usr/bin/python3

import sys
from openpyxl import Workbook, load_workbook
from tqdm import tqdm
import sqlite3
import re

from match import match_a, match_b, match_cd, match_ef, match_g, match_h
from reg import reg_rank, reg_regiment, reg_date, rank_lvl, reg_name, reg_date2

import os.path

from time import sleep

import datetime

import openpyxl

def adapt_date_iso(val):
    """Adapt datetime.date to ISO 8601 date."""
    return val.isoformat()

def adapt_datetime_iso(val):
    """Adapt datetime.datetime to timezone-naive ISO 8601 date."""
    return val.isoformat()

def adapt_datetime_epoch(val):
    """Adapt datetime.datetime to Unix timestamp."""
    return int(val.timestamp())

sqlite3.register_adapter(datetime.date, adapt_date_iso)
sqlite3.register_adapter(datetime.datetime, adapt_datetime_iso)
sqlite3.register_adapter(datetime.datetime, adapt_datetime_epoch)

def convert_date(val):
    """Convert ISO 8601 date to datetime.date object."""
    return datetime.date.fromisoformat(val.decode())

def convert_datetime(val):
    """Convert ISO 8601 datetime to datetime.datetime object."""
    return datetime.datetime.fromisoformat(val.decode())

def convert_timestamp(val):
    """Convert Unix epoch timestamp to datetime.datetime object."""
    return datetime.datetime.fromtimestamp(int(val))

sqlite3.register_converter("date", convert_date)
sqlite3.register_converter("datetime", convert_datetime)
sqlite3.register_converter("timestamp", convert_timestamp)

openpyxl.reader.excel.warnings.simplefilter(action='ignore')

def prettyrow(prow):

    #idkey = prow[0]
    #file = prow[1]
    #sheet = prow[2]
    #row = prow[3]
    #regiment = prow[4]
    #rank = prow[5]
    #name = prow[6]
    #person = prow[7]
    #page = prow[8]
    #deleted = prow[9]
    #handwritten = prow[10]
    #date = prow[11]
    date2 = prow[12]
    #annotations = prow[13]
    reg_regiment = prow[14]
    reg_rank = prow[15]
    rank_lvl = prow[16]
    #surname = prow[17]
    #given = prow[18]
    #middlenames = prow[19]
    #title = prow[20]
    #namesuffix = prow[21]
    #nickname = prow[22]
    #fileyear = prow[23]
    reg_date = prow[24]
    #link_category = prow[25]
    #link_score = prow[26]
    #link_log = prow[27]
    #row_count = prow[28]
    reg_name = prow[29]
    reg_date2 = prow[30]
    reg_date2rank = prow[31]
    reg_date2lvl = prow[32]
    # reg_rank2 = prow[33]
    # rank2_lvl =  prow[34]
    # datealt =  prow[35]
    orig_sheet =  prow[36]

    return ('{:30.30}'.format(str(reg_name))
            + ' ' + '{:2.2}'.format(str(orig_sheet))
            #+ ' ' + '{:20.20}'.format(surname)
            #+ ' ' + '{:10.10}'.format(given)
            #+ ' ' + '{:10.10}'.format(middlenames)
            #+ ' ' + '{:10.10}'.format(title)
            #+ ' ' + '{:10.10}'.format(namesuffix)
            #+ ' ' + '{:10.10}'.format(nickname)
            + ' ' + '{:40.40}'.format(str(reg_regiment))
            + '  ' + '{:20.20}'.format(str(reg_rank))
            + '  ' + '{:5.5}'.format(str(rank_lvl))
            + '  ' + '{:10.10}'.format(str(reg_date))
            #+ '  ' + '{:10.10}'.format(str(datealt)) We do not really know what to do with datealt yet
            + '  ' + '{:20.20}'.format(str(date2))
            + '  ' + '{:10.10}'.format(str(reg_date2))
            + '  ' + '{:20.20}'.format(str(reg_date2rank))
            + '  ' + '{:5.5}'.format(str(reg_date2lvl))
            )

def prettymatch(mmatch):

      print('------ - ', prettyrow(mmatch[0]))

      for match in mmatch[1][:3]:
        print('{:06.3f}'.format(match[2]) # score
          + ' ' + '{:1.1}'.format(str(match[3])) # category
          + '  ' + prettyrow(match[0]) # row
          + ' ' + '{:40.40}'.format(match[4]) # log
          )


def make_link_idkey(aidkey, bidkey):

    return aidkey + '_' + bidkey


def make_link(cur, rowa, tmatch, tmatch_cat, tmatch_log):

    aidkey = rowa[0]
    #afile = rowa[1]
    #asheet = rowa[2]
    #arow = rowa[3]
    #aregiment = rowa[4]
    #arank = rowa[5]
    #aname = rowa[6]
    #aperson = rowa[7]
    #apage = rowa[8]
    #adeleted = rowa[9]
    #ahandwritten = rowa[10]
    #adate = rowa[11]
    #adate2 = rowa[12]
    #aannotations = rowa[13]
    #areg_regiment = rowa[14]
    #areg_rank = rowa[15]
    #arank_lvl = rowa[16]
    #asurname = rowa[17]
    #agiven = rowa[18]
    #amiddlenames = rowa[19]
    #atitle = rowa[20]
    #anamesuffix = rowa[21]
    #anickname = rowa[22]
    #afileyear = rowa[23]
    #areg_date = rowa[24]
    #alink_category = rowa[25]
    #alink_score = rowa[26]
    #alink_log = rowa[27]
    #arow_count = rowa[28] 
    #areg_name = rowa[29] 

    link_score = tmatch[2]

    #link_category = tmatch[3]
    link_category = tmatch_cat

    #link_log = tmatch[4]
    link_log = tmatch_log

    rowb = tmatch[0]

    bidkey = rowb[0]
    #bfile = rowb[1]
    #bsheet = rowb[2]
    #brow = rowb[3]
    #bregiment = rowb[4]
    #brank = rowb[5]
    #bname = rowb[6]
    #bperson = rowb[7]
    #bpage = rowb[8]
    #bdeleted = rowb[9]
    #bhandwritten = rowb[10]
    #bdate = rowb[11]
    #bdate2 = rowb[12]
    #bannotations = rowb[13]
    #breg_regiment = rowb[14]
    #breg_rank = rowb[15]
    #brank_lvl = rowb[16]
    #bsurname = rowb[17]
    #bgiven = rowb[18]
    #bmiddlenames = rowb[19]
    #btitle = rowb[20]
    #bnamesuffix = rowb[21]
    #bnickname = rowb[22]
    #bfileyear = rowb[23]
    #breg_date = rowb[24]
    #blink_category = rowb[25]
    #blink_score = rowb[26]
    #blink_log = rowb[27]
    #brow_count = rowb[28]
    #breg_name = rowb[29]


    idkey = make_link_idkey(aidkey, bidkey)
    
    #print('make link: ' + str(idkey) + ' ' + str(link_category) + ' ' + str(link_score) + ' ' + str(link_log))
    
    while True:
        try:
            cur.execute('INSERT INTO link VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?);', ([idkey, aidkey, bidkey, link_category, link_score, link_log, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None]))   
        except Exception as error:
            print("DB Problem:", error, idkey, basename, sheet, row)
            sleep(2)
            continue
        else:
            break

def failed_link(cur, rowa, tmatch, tmatch_cat, tmatch_log):

    aidkey = rowa[0]
    #afile = rowa[1]
    #asheet = rowa[2]
    #arow = rowa[3]
    #aregiment = rowa[4]
    #arank = rowa[5]
    #aname = rowa[6]
    #aperson = rowa[7]
    #apage = rowa[8]
    #adeleted = rowa[9]
    #ahandwritten = rowa[10]
    #adate = rowa[11]
    #adate2 = rowa[12]
    #aannotations = rowa[13]
    #areg_regiment = rowa[14]
    #areg_rank = rowa[15]
    #arank_lvl = rowa[16]
    #asurname = rowa[17]
    #agiven = rowa[18]
    #amiddlenames = rowa[19]
    #atitle = rowa[20]
    #anamesuffix = rowa[21]
    #anickname = rowa[22]
    #afileyear = rowa[23]
    #areg_date = rowa[24]
    #alink_category = rowa[25]
    #alink_score = rowa[26]
    #alink_log = rowa[27]
    #arow_count = rowa[28] 
    #areg_name = rowa[29]

    #aorig_sheet = rowa[36]
    #aorig_idkey = rowa[37]

    link_score = tmatch[2]

    #link_category = tmatch[3]
    link_category = tmatch_cat

    #link_log = tmatch[4]
    link_log = tmatch_log

    rowb = tmatch[0]

    bidkey = rowb[0]
    #bfile = rowb[1]
    #bsheet = rowb[2]
    #brow = rowb[3]
    #bregiment = rowb[4]
    #brank = rowb[5]
    #bname = rowb[6]
    #bperson = rowb[7]
    #bpage = rowb[8]
    #bdeleted = rowb[9]
    #bhandwritten = rowb[10]
    #bdate = rowb[11]
    #bdate2 = rowb[12]
    #bannotations = rowb[13]
    #breg_regiment = rowb[14]
    #breg_rank = rowb[15]
    #brank_lvl = rowb[16]
    #bsurname = rowb[17]
    #bgiven = rowb[18]
    #bmiddlenames = rowb[19]
    #btitle = rowb[20]
    #bnamesuffix = rowb[21]
    #bnickname = rowb[22]
    #bfileyear = rowb[23]
    #breg_date = rowb[24]
    #blink_category = rowb[25]
    #blink_score = rowb[26]
    #blink_log = rowb[27]
    #brow_count = rowb[28]
    #breg_name = rowb[29]

    #borig_sheet = rowb[36]
    #borig_idkey = rowb[37]

    idkey = make_link_idkey(aidkey, bidkey)

    #print('failed_link: ' + str(idkey) + ' ' + str(link_category) + ' ' + str(link_score) + ' ' + str(link_log))
    while True:
        try:
            cur.execute('INSERT INTO link VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?);', ([idkey, aidkey, bidkey, link_category, link_score, link_log, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None]))
        except Exception as error:
            print("DB Problem:", error, idkey, basename, sheet, row)
            sleep(2)
            continue
        else:
            break

def make_row_idkey(file, sheet, row):

    #return file[36:41] + '-' + "{:02d}".format(int(sheet[0:2].replace('.', ''))) + '-' + "{:06d}".format(row)
    return file[3:8] + '-' + "{:02d}".format(int(sheet[0:2].replace('.', ''))) + '-' + "{:06d}".format(row)

def create_link_table(cur):

    cur.execute('CREATE TABLE IF NOT EXISTS link(idkey PRIMARY KEY, aidkey, bidkey, link_category, link_score, link_log, afile, bfile, asheet, bsheet, arow, brow, areg_name, breg_name, areg_regiment, breg_regiment, areg_rank, breg_rank, areg_date, breg_date, adate2, bdate2, aannotations, bannotations, adeleted, bdeleted, ahandwritten, bhandwritten, afileyear, bfileyear, asurname, bsurname, agiven, bgiven, arank_lvl, brank_lvl, aorig_sheet, borig_sheet, aorig_idkey, borig_idkey, apage, bpage, aname, bname);')

command = sys.argv[1]

if (command == 'ingest' or command == 'ingest-test'):

    file = sys.argv[2]
    wb = load_workbook(filename = file, read_only=True) # If you load and save an xlsx file from libreoffice, it writes the length of each sheet so the optimised reader knows the sheet length.
    #wb = load_workbook(filename = file) # On files saved out of Excel, sometimes the length of each sheet is unknown with the optimised reader, so you have to use the unoptimised reader.

    sheet = sys.argv[3]
    ws = wb[sheet]

    dbfile = sys.argv[4]
    dbsheet = sys.argv[5]

    con = sqlite3.connect('british-army-officers.db')
    cur = con.cursor()
    # ALTER TABLE row ADD COLUMN typev;
    # ALTER TABLE row ADD COLUMN reduced;
    # ALTER TABLE row ADD COLUMN datealt;
    # ALTER TABLE row ADD COLUMN dategen;
    # ALTER TABLE row ADD COLUMN datelgen;
    # ALTER TABLE row ADD COLUMN datemgen;
    # ALTER TABLE row ADD COLUMN datecol;
    # ALTER TABLE row ADD COLUMN datelcol;
    # ALTER TABLE row ADD COLUMN datemaj;
    # ALTER TABLE row ADD COLUMN datefm;
    # ALTER TABLE row ADD COLUMN reg_date2;
    # ALTER TABLE row ADD COLUMN reg_date2rank;
    # ALTER TABLE row ADD COLUMN reg_date2lvl;
    # ALTER TABLE row ADD COLUMN reg_rank2;
    # ALTER TABLE row ADD COLUMN rank2_lvl;
    # ALTER TABLE row ADD COLUMN orig_sheet;
    # ALTER TABLE row ADD COLUMN orig_idkey;
    # ALTER TABLE row ADD COLUMN name_rank;
    # ALTER TABLE row ADD COLUMN name_rank_lvl;

    cur.execute('CREATE TABLE IF NOT EXISTS row(idkey PRIMARY KEY, file, sheet, row, regiment, rank, name, person, page, deleted, handwritten, date, date2, annotations, reg_regiment, reg_rank, rank_lvl, surname, given, middlenames, title, namesuffix, nickname, fileyear, reg_date, link_category, link_score, link_log, row_count, reg_name, linked_to_file, linked_to_sheet, linked_to_row, linked_to_rowid, typev, reduced, datealt, companies, place, rank2, untitled, totheking, totheprince, rankunlisted, lv, sv, dv, othertext, role, medal, foreignorder, department, datefrom, dateto, regnumber, description, dategen, datelgen, datemgen, datecol, datelcol, datemaj, datefm, reg_date2, reg_date2rank, reg_date2lvl, reg_rank2, rank2_lvl, orig_sheet, orig_idkey, name_rank, name_rank_lvl);')
   
    #print(ws.max_row)

    for i, xlrow in enumerate(tqdm(ws.iter_rows(min_row=2), total=ws.max_row, miniters=10, ascii=True, ncols=60)):

      row = i + 1

      if len(xlrow) > 3:

          date = None
          date2 = None
          regiment = None
          annotations = None
          rank = None
          name = None

          companies = None
          place = None

          rank2 = None

          typev = None
          reduced = None
          datealt = None

          untitled = None
          totheking = None
          totheprince = None
          rankunlisted = None
          lv = None
          sv = None
          dv = None
          othertext = None
          role = None
          medal = None
          foreignorder = None
          department = None
          datefrom = None
          dateto = None
          regnumber = None
          description = None

          dategen = None
          datelgen = None
          datemgen = None
          datecol = None
          datelcol = None
          datemaj = None
          datefm = None

          deleted = None

          #print(len(xlrow))

          # Hacks for Fencibles
          #if file == 'data/drive/WO65 transcribed data/Fencibles Army Lists transcribed data/1793 Fencibles - PRINTED.xlsx' and sheet == '10. Regiments': file = 'data/drive/WO65 transcribed data/WO 65-43 Army List 1793.xlsx'; sheet = '55. Fencibles'
          #if file == 'data/drive/WO65 transcribed data/Fencibles Army Lists transcribed data/1794 Fencible List - HANDWRITTEN.xlsx' and sheet == '10. Regiments': file = 'data/drive/WO65 transcribed data/WO 65-44 Army List 1794.xlsx'; sheet = '56. Fencibles Handwritten'
          #if file == 'data/drive/WO65 transcribed data/Fencibles Army Lists transcribed data/1794 Fencibles - PRINTED.xlsx' and sheet == '10. Regiments': file = 'data/drive/WO65 transcribed data/WO 65-44 Army List 1794.xlsx'; sheet = '55. Fencibles'
          #if file == 'data/drive/WO65 transcribed data/Fencibles Army Lists transcribed data/1795 Fencibles - PRINTED.xlsx' and sheet == '10. Regiments': file = 'data/drive/WO65 transcribed data/WO 65-45 Army List 1795.xlsx'; sheet = '55. Fencibles'
          #if file == 'data/drive/WO65 transcribed data/Fencibles Army Lists transcribed data/1796 Fencibles - PRINTED.xlsx' and sheet == '10. Regiments': file = 'data/drive/WO65 transcribed data/WO 65-46 Army List 1796.xlsx'; sheet = '55. Fencibles'
          #if file == 'data/drive/WO65 transcribed data/Fencibles Army Lists transcribed data/1797 Fencibles - PRINTED.xlsx' and sheet == '10. Regiments': file = 'data/drive/WO65 transcribed data/WO 65-47 Army List 1797.xlsx'; sheet = '55. Fencibles'
          #if file == 'data/drive/WO65 transcribed data/Fencibles Army Lists transcribed data/1799 Fencibles - PRINTED.xlsx' and sheet == '10. Regiments': file = 'data/drive/WO65 transcribed data/WO 65-49 Army List 1799.xlsx'; sheet = '55. Fencibles'
          #if file == 'data/drive/WO65 transcribed data/Fencibles Army Lists transcribed data/1800 Fencibles - PRINTED.xlsx' and sheet == '10. Regiments': file = 'data/drive/WO65 transcribed data/WO 65-50 Army List 1800.xlsx'; sheet = '55. Fencibles'
          #if file == 'data/drive/WO65 transcribed data/Fencibles Army Lists transcribed data/1801 Fencibles - PRINTED.xlsx' and sheet == '10. Regiments': file = 'data/drive/WO65 transcribed data/WO 65-51 Army List 1801.xlsx'; sheet = '55. Fencibles'
          # End of hacks for fensibles

          if dbsheet == '1. General and Field Officers':

            name = xlrow[4].value
            rank = xlrow[3].value
            #regiment = xlrow[3].value
            page = xlrow[0].value
            deleted = xlrow[1].value
            handwritten = xlrow[2].value
            date = xlrow[5].value
            #date2 = xlrow[7].value
            annotations = xlrow[14].value

            #untitled = xlrow[13].value
            regiment = xlrow[13].value

            dategen = xlrow[6].value
            datelgen = xlrow[7].value
            datemgen = xlrow[8].value
            datecol = xlrow[9].value
            datelcol = xlrow[10].value
            datemaj = xlrow[11].value
            datefm = xlrow[12].value

          elif dbsheet == '2. Aides-de-Camp':

              name = xlrow[5].value
              rank = xlrow[4].value

              if len(xlrow) > 8:
                regiment = xlrow[8].value

              page = xlrow[0].value
              deleted = xlrow[1].value
              handwritten = xlrow[2].value
              date = xlrow[6].value
              date2 = xlrow[7].value

              if len(xlrow) > 9:
                annotations = xlrow[9].value

              untitled = xlrow[3].value

          elif dbsheet == '3. Local Rank':

            name = xlrow[4].value
            rank = xlrow[3].value
            #regiment = xlrow[3].value
            page = xlrow[0].value
            deleted = xlrow[1].value
            handwritten = xlrow[2].value
            date = xlrow[5].value
            #date2 = xlrow[7].value
            #annotations = xlrow[8].value

            place = xlrow[6].value

          elif dbsheet == '4. Staff & Miscellaneous':

            if len(xlrow) > 6:

                page = xlrow[0].value
                deleted = xlrow[1].value
                handwritten = xlrow[2].value
                place = xlrow[3].value
                rank = xlrow[5].value
                name = xlrow[6].value
                date = xlrow[7].value
                date2 = xlrow[8].value
                regiment = xlrow[9].value
                annotations = xlrow[10].value

                rankunlisted = xlrow[5].value

                #lv = xlrow[10].value
                #sv = xlrow[11].value
                #dv = xlrow[12].value
                #othertext = xlrow[13].value

          elif dbsheet == '5. Staff Officers':

            name = xlrow[6].value
            rank = xlrow[5].value
            #regiment = xlrow[3].value
            page = xlrow[0].value
            deleted = xlrow[1].value
            handwritten = xlrow[2].value
            date = xlrow[7].value
            date2 = xlrow[8].value

            if len(xlrow) > 9:
                annotations = xlrow[9].value

            place = xlrow[4].value

            role = xlrow[3].value

          elif dbsheet == '6. Honorary Distinctions': # Begins in 1813

            #if xlrow[6].value != None and xlrow[4].value != None:
            #    name = xlrow[6].value + ' ' + xlrow[4].value
            #elif xlrow[4].value != None:
            #    name = xlrow[4].value
            #elif xlrow[6].value != None:
            #    name = xlrow[6].value

            if len(xlrow) > 5:

                name = xlrow[5].value

                rank = xlrow[4].value
                regiment = xlrow[6].value
                page = xlrow[0].value
                deleted = xlrow[1].value
                handwritten = xlrow[2].value

                ##date = xlrow[6].value
                ##date2 = xlrow[7].value

                if len(xlrow) > 8:

                    annotations = xlrow[8].value

                medal = xlrow[3].value
                rank2 = xlrow[7].value


          elif dbsheet == '7. Order of the Bath': # Begins in 1815

            if len(xlrow) > 5:

                page = xlrow[0].value
                deleted = xlrow[1].value
                handwritten = xlrow[2].value
                othertext = xlrow[3].value
                name = xlrow[4].value
                regiment = xlrow[5].value

                if len(xlrow) > 6:

                    annotations = xlrow[6].value

          elif dbsheet == '8. Guelphic Order':

            if len(xlrow) > 4:

                name = xlrow[4].value
                rank = xlrow[3].value
                regiment = xlrow[5].value
                page = xlrow[0].value
                deleted = xlrow[1].value
                handwritten = xlrow[2].value
                #date = xlrow[6].value
                #date2 = xlrow[7].value
                annotations = xlrow[6].value

          elif dbsheet == '9. Foreign Orders':

            if len(xlrow) > 3:

                name = xlrow[3].value
                #rank = xlrow[3].value

                if len(xlrow) > 4:
                    regiment = xlrow[4].value

                page = xlrow[0].value
                deleted = xlrow[1].value
                handwritten = xlrow[2].value
                #date = xlrow[6].value
                #date2 = xlrow[7].value

                if len(xlrow) > 6:
                    annotations = xlrow[6].value

                if len(xlrow) > 5:
                    foreignorder = xlrow[5].value

          elif dbsheet == '10. Regiments':

            name = xlrow[5].value
            rank = xlrow[4].value
            regiment = xlrow[3].value
            page = xlrow[0].value
            deleted = xlrow[1].value
            handwritten = xlrow[2].value
            date = xlrow[6].value
            date2 = xlrow[7].value
            annotations = xlrow[8].value

          elif dbsheet == '11. Independent Companies':

            if len(xlrow) > 6:

                name = xlrow[6].value
                rank = xlrow[5].value
                #regiment = xlrow[3].value
                page = xlrow[0].value
                deleted = xlrow[1].value
                handwritten = xlrow[2].value
                date = xlrow[7].value

                if len(xlrow) > 8:
                    date2 = xlrow[8].value

                if len(xlrow) > 9:
                    annotations = xlrow[9].value

                companies = xlrow[3].value
                place = xlrow[4].value

          elif dbsheet == '12. Invalids':

            name = xlrow[6].value
            rank = xlrow[5].value
            #regiment = xlrow[3].value
            page = xlrow[0].value
            deleted = xlrow[1].value
            handwritten = xlrow[2].value
            date = xlrow[7].value
            date2 = xlrow[8].value
            annotations = xlrow[9].value

            companies = xlrow[3].value
            #place = xlrow[4].value
            regiment = xlrow[4].value

          elif dbsheet == '13. Garrisons':

            name = xlrow[5].value
            rank = xlrow[4].value
            #regiment = xlrow[3].value
            page = xlrow[0].value
            deleted = xlrow[1].value
            handwritten = xlrow[2].value
            #date = xlrow[6].value
            #date2 = xlrow[7].value
            annotations = xlrow[9].value

            #place = xlrow[3].value
            regiment = xlrow[3].value

            lv = xlrow[6].value
            sv = xlrow[7].value
            dv = xlrow[8].value

          elif dbsheet == '14. Royal Artillery etc':

            name = xlrow[5].value
            rank = xlrow[4].value
            regiment = xlrow[3].value
            page = xlrow[0].value
            deleted = xlrow[1].value
            handwritten = xlrow[2].value
            date = xlrow[6].value
            date2 = xlrow[7].value
            annotations = xlrow[8].value

          elif dbsheet == '15. Marines':

            name = xlrow[5].value
            rank = xlrow[4].value
            regiment = xlrow[3].value
            page = xlrow[0].value
            deleted = xlrow[1].value
            handwritten = xlrow[2].value
            date = xlrow[6].value
            date2 = xlrow[8].value
            annotations = xlrow[9].value

            rank2 = xlrow[7].value

          elif dbsheet == '16. Officers on Full Pay':

            if len(xlrow) > 4:
                name = xlrow[4].value
                rank = xlrow[3].value
                #regiment = xlrow[3].value
                page = xlrow[0].value
                deleted = xlrow[1].value
                handwritten = xlrow[2].value

                if len(xlrow) > 5:
                    date = xlrow[5].value

                if len(xlrow) > 7:
                    date2 = xlrow[7].value

                if len(xlrow) > 8:
                    annotations = xlrow[8].value

                if len(xlrow) > 6:
                    rank2 = xlrow[6].value

          elif dbsheet == '17-21. Military Departments':

            name = xlrow[5].value
            rank = xlrow[4].value
            regiment = xlrow[1].value
            page = xlrow[0].value
            deleted = xlrow[2].value
            handwritten = xlrow[3].value

            if len(xlrow) > 6:
                date = xlrow[6].value

            #date2 = xlrow[7].value

            if len(xlrow) > 7:
                annotations = xlrow[7].value

            #department = xlrow[1].value

          elif dbsheet == '22. Reduced Corps':

            name = xlrow[5].value
            rank = xlrow[4].value
            regiment = xlrow[3].value
            page = xlrow[0].value
            deleted = xlrow[1].value
            handwritten = xlrow[2].value

            if len(xlrow) > 6:
                date = xlrow[6].value

            if len(xlrow) > 8:
                date2 = xlrow[8].value

            if len(xlrow) > 9:
                annotations = xlrow[9].value

            if len(xlrow) > 7:
                rank2 = xlrow[7].value

          elif dbsheet == '23. Companies of Foot':

            if len(xlrow) > 7:
                name = xlrow[7].value
                rank = xlrow[6].value
                regiment = xlrow[5].value
                page = xlrow[0].value
                deleted = xlrow[1].value
                handwritten = xlrow[2].value
                #date = xlrow[6].value
                #date2 = xlrow[7].value
                annotations = xlrow[8].value

                datefrom = xlrow[3].value
                dateto = xlrow[4].value

          elif dbsheet == '24. Companies of Invalids':

            name = xlrow[5].value
            rank = xlrow[4].value
            #regiment = xlrow[3].value
            page = xlrow[0].value
            deleted = xlrow[1].value
            handwritten = xlrow[2].value
            date = xlrow[6].value
            date2 = xlrow[7].value
            annotations = xlrow[8].value

            place = xlrow[3].value

          elif dbsheet == '25. Officers Unattached':

            name = xlrow[4].value
            rank = xlrow[3].value
            #regiment = xlrow[3].value
            page = xlrow[0].value
            deleted = xlrow[1].value
            handwritten = xlrow[2].value
            date = xlrow[5].value
            date2 = xlrow[6].value
            annotations = xlrow[7].value

          elif dbsheet == '26. Foot Guards':

            name = xlrow[5].value
            rank = xlrow[4].value
            #regiment = xlrow[3].value
            page = xlrow[0].value
            deleted = xlrow[1].value
            handwritten = xlrow[2].value
            date = xlrow[6].value
            date2 = xlrow[7].value
            annotations = xlrow[8].value

            place = xlrow[3].value

          elif dbsheet == '27. Retired & Reduced':

            name = xlrow[5].value
            rank = xlrow[4].value
            regiment = xlrow[3].value
            page = xlrow[0].value
            deleted = xlrow[1].value
            handwritten = xlrow[2].value
            date = xlrow[6].value
            date2 = xlrow[8].value
            annotations = xlrow[9].value

            rank2 = xlrow[7].value

          elif dbsheet == '28-31. Officers on Half Pay':

            if len(xlrow) > 8:
                name = xlrow[8].value
                rank = xlrow[7].value
                regiment = xlrow[6].value
                page = xlrow[0].value
                deleted = xlrow[2].value
                handwritten = xlrow[3].value
                #date = xlrow[5].value # Too different to date on other sheets to be treated the same way
                #date2 = xlrow[7].value

                if len(xlrow) > 9:
                    annotations = xlrow[9].value

                typev = xlrow[1].value
                reduced = xlrow[4].value
                datealt = xlrow[5].value # Too different to date to be treated the same way

          elif dbsheet == '32. New Independent Companies':

            name = xlrow[5].value
            rank = xlrow[4].value
            regiment = xlrow[3].value
            page = xlrow[0].value
            deleted = xlrow[1].value
            handwritten = xlrow[2].value
            date = xlrow[6].value
            date2 = xlrow[7].value
            annotations = xlrow[8].value

          elif dbsheet == '33. Succession of Colonels':

            if len(xlrow) > 5:
                name = xlrow[5].value
                #rank = xlrow[4].value
                regiment = xlrow[3].value
                page = xlrow[0].value
                deleted = xlrow[1].value
                handwritten = xlrow[2].value


            if len(xlrow) > 6:
                date = xlrow[6].value
                #date2 = xlrow[7].value

                if len(xlrow) > 7:
                    annotations = xlrow[7].value

                regnumber = xlrow[4].value

          elif dbsheet == '34. Casualties':

            name = xlrow[5].value
            rank = xlrow[4].value
            regiment = xlrow[6].value
            page = xlrow[0].value
            deleted = xlrow[1].value
            handwritten = xlrow[2].value
            #date = xlrow[6].value
            #date2 = xlrow[7].value
            if len(xlrow) > 7:
                annotations = xlrow[7].value

            typev = xlrow[3].value

          elif dbsheet == '35. Alterations While Printing':

            name = xlrow[4].value
            rank = xlrow[5].value
            regiment = xlrow[3].value
            page = xlrow[0].value
            deleted = xlrow[1].value
            handwritten = xlrow[2].value
            date = xlrow[7].value
            #date2 = xlrow[7].value
            annotations = xlrow[8].value

            description = xlrow[6].value

          elif dbsheet == '36. Errata':

            #name = xlrow[5].value
            #rank = xlrow[4].value
            #regiment = xlrow[3].value
            page = xlrow[0].value
            deleted = xlrow[1].value
            handwritten = xlrow[2].value
            #date = xlrow[6].value
            #date2 = xlrow[7].value
            annotations = xlrow[4].value

            description = xlrow[3].value

          elif dbsheet == '55. Fencibles':

            name = xlrow[4].value
            rank = xlrow[3].value
            regiment = xlrow[2].value
            page = xlrow[0].value
            deleted = xlrow[1].value
            handwritten = False

            if (len(xlrow) > 5):
                date = xlrow[5].value

            if (len(xlrow) > 6):
                date2 = xlrow[6].value

            if (len(xlrow) > 7):
                annotations = xlrow[7].value

          elif dbsheet == '56. Fencibles (handwritten)':

            name = xlrow[3].value
            rank = xlrow[2].value
            regiment = xlrow[1].value
            page = xlrow[0].value
            #deleted = xlrow[1].value
            handwritten = True
            date = xlrow[4].value

            if (len(xlrow) > 5):
                date2 = xlrow[5].value

            if (len(xlrow) > 6):
                annotations = xlrow[6].value

          elif dbsheet == '61. General and Field Officers (handwritten)':

            handwritten = True
            page = xlrow[0].value
            deleted = xlrow[1].value
            rank = xlrow[2].value
            name = xlrow[3].value
            date = xlrow[4].value
            datelgen = xlrow[5].value
            datemgen = xlrow[6].value
            date2 = xlrow[7].value
            regiment = xlrow[8].value
            datemaj = xlrow[9].value
            datefm = xlrow[10].value
            untitled = xlrow[11].value
            annotations = xlrow[12].value

          elif dbsheet == '70. Regiments (handwritten)':

            handwritten = True
            page = xlrow[0].value
            deleted = xlrow[1].value
            regiment = xlrow[2].value
            rank = xlrow[3].value
            name = xlrow[4].value
            date = xlrow[6].value
            date2 = xlrow[7].value
            annotations = str(xlrow[5].value) + ' || ' + str(xlrow[8].value) + ' || ' + str(xlrow[9].value)
            
          elif dbsheet == '72. Invalids (handwritten)':

            handwritten = True
            page = xlrow[0].value
            deleted = xlrow[1].value
            companies = xlrow[2].value
            regiment = xlrow[3].value
            rank = xlrow[4].value
            name = xlrow[5].value
            date = xlrow[6].value
            date2 = xlrow[7].value
            annotations = xlrow[8].value

          elif dbsheet == '73. Garrisons (handwritten)':

            handwritten = True
            page = xlrow[0].value
            deleted = xlrow[1].value
            regiment = xlrow[2].value
            rank = xlrow[3].value
            name = xlrow[4].value
            lv = xlrow[5].value
            sv = xlrow[6].value
            dv = xlrow[7].value
            annotations = xlrow[8].value
            date = xlrow[9].value

          elif dbsheet == '92. New Independent Companies (handwritten)':

            handwritten = True
            page = xlrow[0].value
            deleted = xlrow[1].value
            regiment = xlrow[2].value
            rank = xlrow[3].value
            name = xlrow[4].value
            date = xlrow[5].value
            date2 = xlrow[6].value
            annotations = xlrow[7].value

          elif dbsheet == '71. Independent Companies (handwritten)':

            handwritten = True
            page = xlrow[0].value
            deleted = xlrow[1].value
            companies = xlrow[2].value
            regiment = xlrow[3].value
            rank = xlrow[4].value

            if len(xlrow) > 5:

                name = xlrow[5].value
                date = xlrow[6].value
                date2 = xlrow[7].value
                annotations = xlrow[8].value

          elif dbsheet == '74. Royal Artillery etc (handwritten)':

            handwritten = True
            page = xlrow[0].value
            deleted = xlrow[1].value
            regiment = xlrow[2].value
            rank = xlrow[3].value

            if len(xlrow) > 4:

                name = xlrow[4].value
                date = xlrow[5].value
                date2 = xlrow[6].value
                annotations = xlrow[7].value

          elif dbsheet == '76. Officers on Full Pay (handwritten)':

            handwritten = True
            page = xlrow[0].value
            deleted = xlrow[1].value
            rank = xlrow[2].value
            name = xlrow[3].value
            date = xlrow[4].value
            rank2 = xlrow[5].value

            if len(xlrow) > 6:

                date2 = xlrow[6].value
                annotations = xlrow[7].value

          elif dbsheet == '82. Reduced Corps (handwritten)':

            handwritten = True
            page = xlrow[0].value
            deleted = xlrow[1].value
            regiment = xlrow[2].value
            rank = xlrow[3].value
            name = xlrow[4].value
            date = xlrow[5].value
            rank2 = xlrow[6].value

            if len(xlrow) > 7:

                date2 = xlrow[7].value
                annotations = xlrow[8].value

          elif dbsheet == '88-91. Officers on Half Pay (handwritten)':

            handwritten = True
            page = xlrow[0].value
            deleted = xlrow[1].value
            typev = xlrow[2].value
            reduced = xlrow[3].value
            datealt = xlrow[4].value # Too different to date to be treated the same way
            regiment = xlrow[5].value
            rank = xlrow[6].value

            if len(xlrow) > 7:

                name = xlrow[7].value
                annotations = xlrow[8].value

          elif dbsheet == '63. Local Rank (handwritten)':

            handwritten = True
            page = xlrow[0].value
            deleted = xlrow[1].value
            rank = xlrow[2].value
            name = xlrow[3].value
            date = xlrow[4].value
            place = xlrow[5].value
            annotations = xlrow[6].value

          else:

              print()
              print()
              print('DB sheet name "' + dbsheet + '" was not recognised for processing.')
              print()
              exit(1)

          reg_name_value = None
          surname = None
          given = None
          middlenames = None
          title = None
          namesuffix = None
          nickname = None
          name_rank = None
          name_rank_lvl = None
        
          #if name != None:

          #  reg_name_value, surname, given, middlenames, title, namesuffix, nickname, name_rank = reg_name(name)

          #if name_rank != '':

          #    name_rank_lvl = rank_lvl(name_rank)


          reg_date2_value = None
          reg_date2rank_value = None
          reg_date2lvl_value = None

          #if date2 != None:

          #    reg_date2_value, reg_date2rank_value, reg_date2lvl_value = reg_date2(date2)


          reg_rank2 = None
          rank2_lvl = None

          #if rank2 != None:

          #    reg_rank2 = reg_rank(rank2)
          #    rank2_lvl = rank_lvl(reg_rank2)

          #basename = os.path.basename(file)[:-5]
          fileyear = int(dbfile[19:])

          person = None
          link_category = None
          link_score = None
          link_log = None
          row_count = 1

          linked_to_file = None
          linked_to_sheet = None
          linked_to_row = None
          linked_to_rowid = None
          orig_sheet = None
          orig_idkey = None

          reg_regiment = None
          reg_rank = None
          rank_lvl = None
          reg_date = None

          if (name is not None):
           
            idkey = make_row_idkey(dbfile, dbsheet, row)

            if command == 'ingest-test':

                print(idkey + '|NAME:' + name + '|REGIMENT:' + str(regiment) + '|RANK:' + str(rank) + '|RANK2:' + str(rank2) + '|DATE:' + str(date) + '|DATE2:' + str(date2) + '|PAGE:' + str(page) + '|DELETED:' + str(deleted) + '|HANDWRITTEN:' + str(handwritten) + '|ANNOTATIONS:' + str(annotations) + '|UNTITLED:' + str(untitled))

            # The check on whether records had already been ingested was too slow.
            # Quicker to do:
            # SELECT COUNT(*) FROM row WHERE file = 'WO 65-56 Army List 1806';
            # DELETE FROM row WHERE file = 'WO 65-56 Army List 1806';
            #
            #print('Checking if record is already indexed')
            #cur.execute('SELECT 1 FROM row WHERE file = ? and sheet = ? and row = ? LIMIT 1;', ([basename, sheet, row]))
            #res = cur.fetchone()
            #if (res is None):
            while True:
                try:
                    cur.execute('INSERT INTO row VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?);', ([idkey, dbfile, dbsheet, row, regiment, rank, name, person, page, deleted, handwritten, date, date2, annotations, reg_regiment, reg_rank, rank_lvl, surname, given, middlenames, title, namesuffix, nickname, fileyear, reg_date, link_category, link_score, link_log, row_count, reg_name_value, linked_to_file, linked_to_sheet, linked_to_row, linked_to_rowid, typev, reduced, datealt, companies, place, rank2, untitled, totheking, totheprince, rankunlisted, lv, sv, dv, othertext, role, medal, foreignorder, department, datefrom, dateto, regnumber, description, dategen, datelgen, datemgen, datecol, datelcol, datemaj, datefm, reg_date2_value, reg_date2rank_value, reg_date2lvl_value, reg_rank2, rank2_lvl, orig_sheet, orig_idkey, name_rank, name_rank_lvl]))
                except Exception as error:
                    print("DB Problem:", error, idkey, dbfile, sheet, row)
                    sleep(2)
                    continue
                else:
                    break

    if command == 'ingest':

      con.commit()
      con.close()

if (command == 'link' or command == 'link-test'):

    filea = sys.argv[2]
    sheeta = sys.argv[3]

    fileb = sys.argv[4]
    sheetb = sys.argv[5]

    logictree = sys.argv[6]

    testmax = None
    if len(sys.argv) == 8: testmax = int(sys.argv[7])

    con = sqlite3.connect('british-army-officers.db')
    cur = con.cursor()

    create_link_table(cur)

    # Get Rows A
    if logictree == 'D' or logictree == 'F':

        # sheet 1 to sheets 14 and 15, where there isn’t a match between a name in sheet 1 and sheet 10.
        cur.execute('SELECT idkey, file, sheet, row, regiment, rank, name, person, page, deleted, handwritten, date, date2, annotations, reg_regiment, reg_rank, rank_lvl, surname, given, middlenames, title, namesuffix, nickname, fileyear, reg_date, link_category, link_score, link_log, row_count, reg_name, reg_date2, reg_date2rank, reg_date2lvl, reg_rank2, rank2_lvl, datealt, orig_sheet, orig_idkey, name_rank, name_rank_lvl, rank2 FROM row WHERE file = ? AND sheet = ? AND person IS NULL;', [filea, sheeta])

    else:

        cur.execute('SELECT idkey, file, sheet, row, regiment, rank, name, person, page, deleted, handwritten, date, date2, annotations, reg_regiment, reg_rank, rank_lvl, surname, given, middlenames, title, namesuffix, nickname, fileyear, reg_date, link_category, link_score, link_log, row_count, reg_name, reg_date2, reg_date2rank, reg_date2lvl, reg_rank2, rank2_lvl, datealt, orig_sheet, orig_idkey, name_rank, name_rank_lvl, rank2 FROM row WHERE file = ? AND sheet = ?;', [filea, sheeta])

    rowsa = cur.fetchall()

    # Get Rows B
    if logictree == 'C':

        # Sheet 10.current rank should be captain or higher in our list of ranks
        cur.execute('SELECT idkey, file, sheet, row, regiment, rank, name, person, page, deleted, handwritten, date, date2, annotations, reg_regiment, reg_rank, rank_lvl, surname, given, middlenames, title, namesuffix, nickname, fileyear, reg_date, link_category, link_score, link_log, row_count, reg_name, reg_date2, reg_date2rank, reg_date2lvl, reg_rank2, rank2_lvl, datealt, orig_sheet, orig_idkey, name_rank, name_rank_lvl, rank2 FROM row WHERE file = ? AND sheet = ? AND rank_lvl >= 7;', ([fileb, sheetb]))

    else:

        cur.execute('SELECT idkey, file, sheet, row, regiment, rank, name, person, page, deleted, handwritten, date, date2, annotations, reg_regiment, reg_rank, rank_lvl, surname, given, middlenames, title, namesuffix, nickname, fileyear, reg_date, link_category, link_score, link_log, row_count, reg_name, reg_date2, reg_date2rank, reg_date2lvl, reg_rank2, rank2_lvl, datealt, orig_sheet, orig_idkey, name_rank, name_rank_lvl, rank2 FROM row WHERE file = ? AND sheet = ?;', ([fileb, sheetb]))

    rowsb = cur.fetchall()

    print('Matching ' + str(len(rowsa)) + ' rows (A) against ' + str(len(rowsb)) + ' rows (B).')

    if len(rowsa) > 0 and len(rowsb) > 0:
 
        matched = {}
        failed = []
        #newapp = []

        for i, rowa in enumerate(tqdm(rowsa, total=len(rowsa), miniters=1, ascii=True, ncols=60)):

          if testmax != None and i >= testmax:
            break

          if logictree == 'A': # Sheet 10 to Sheet 10 (previous year) logic

              matches = match_a(rowa, rowsb)

          elif logictree == 'B': # Name only

              matches = match_b(rowa, rowsb)

          elif logictree == 'C': # Sheet 1 to Sheet 10 (same year) logic

              matches = match_cd(rowa, rowsb)

          elif logictree == 'D': # Sheet 1 to Sheet 14 (same year) logic

              matches = match_cd(rowa, rowsb)

          elif logictree == 'E': # Sheet 34 to Sheet 10 (same year) logic

              matches = match_ef(rowa, rowsb)

          elif logictree == 'F': # Sheet 34 to Sheet 14 (same year) logic

              matches = match_ef(rowa, rowsb)

          elif logictree == 'G': # Sheet 28 to Sheet 10 (previous year) logic

              cur.execute('SELECT common FROM common_given WHERE term = ?;', ([rowa[18]]))
              common_given = cur.fetchone()[0]
              cur.execute('SELECT common FROM common_surname WHERE term = ?;', ([rowa[17]]))
              common_surname = cur.fetchone()[0]

              matches = match_g(rowa, rowsb, common_given, common_surname)

          elif logictree == 'H': # Handwritten to printed logic

              matches = match_h(rowa, rowsb)

          else:

              print()
              print('Did not recognise logictree: "' + logictree + '"')
              print()
              exit(1)

          matches = sorted(matches, key=lambda x: x[2], reverse=True)

          fmatch = matches[0]
          smatch = matches[1]
          tmatch = matches[2]

          match_cat = fmatch[3]

          if match_cat < 4:

              bidkey = fmatch[0][0]

              if bidkey in matched:
                  #print()
                  #print('Contested match: ' + str(rowa[0]) + ' --> '  + str(browid))
                  #print()
                  matched[bidkey].append((rowa, fmatch, smatch, tmatch))
                  #exit(1)
              else:
                  matched[bidkey] = [(rowa, fmatch, smatch, tmatch)]

          else:

                  failed.append((rowa, fmatch, smatch, tmatch))

        print()
        print('processing matched ...')
        if command == 'link-test':

            categories = []
            categories.append([])
            categories.append([])
            categories.append([])
            categories.append([])
            categories.append([])
            categories.append([])
            categories.append([])
            categories.append([])

        for matchkey in tqdm(matched.keys(), total=len(matched.keys()), miniters=1, ascii=True, ncols=60):

            #print(matchkey)
            matchlist = matched[matchkey]

            if len(matchlist) == 1:

                # just one match

                if command == 'link-test':

                    match_cat = matchlist[0][1][3]
                    categories[match_cat].append((matchlist[0][0], [matchlist[0][1], matchlist[0][2], matchlist[0][3]]))

                elif command == 'link':

                    match_cat = matchlist[0][1][3]
                    match_log = matchlist[0][1][4]
                    make_link(cur, matchlist[0][0], matchlist[0][1], match_cat, match_log)

            else:

                # contested match
                matchlist = sorted(matchlist, key=lambda x: x[1][2], reverse=True)

                if command == 'link-test':

                    match_cat = matchlist[0][1][3]
                    categories[match_cat].append((matchlist[0][0], [matchlist[0][1], matchlist[0][2], matchlist[0][3]]))

                    # Records which missed out become link_category 7
                    for rejmatch in matchlist[1:]:
                        # Tuples are immutable so we need to make a new, fake top hit
                        fakehit = (rejmatch[1][0], rejmatch[1][1], rejmatch[1][2], rejmatch[1][3], 'BETTER: ' + matchlist[0][0][0] + ' ' + matchlist[0][0][6])
                        categories[7].append((rejmatch[0], [fakehit, rejmatch[2], rejmatch[3]]))

                elif command == 'link':

                    match_cat = matchlist[0][1][3]
                    match_log = matchlist[0][1][4]
                    make_link(cur, matchlist[0][0], matchlist[0][1], match_cat, match_log)

                    # Records which missed out become link_category 7
                    for rejmatch in matchlist[1:]:
                        #make_link(cur, rejmatch[0], rejmatch[1], 7, 'BETTER: ' + matchlist[0][0][0] + ' ' + matchlist[0][0][6])
                        failed_link(cur, rejmatch[0], rejmatch[1], 7, 'BETTER: ' + matchlist[0][0][0] + ' ' + matchlist[0][0][6])

        print()
        print('processing failures ..')
        for failedlink in tqdm(failed, total=len(failed), miniters=1, ascii=True, ncols=60):
            
            if command == 'link-test':

                categories[4].append((failedlink[0], [failedlink[1], failedlink[2], failedlink[3]]))

            elif command == 'link':

                match_log = failedlink[1][4]
                failed_link(cur, failedlink[0], failedlink[1], 4, match_log)


        #print()
        #print('processing new appointments ..')
        #for newlink in tqdm(newapp, total=len(newapp), miniters=1, ascii=True, ncols=60):
        #    
        #    if command == 'link-test':
        #
        #        categories[5].append((newlink[0], [newlink[1], newlink[2], newlink[3]]))
        #
        #    elif command == 'link':
        #
        #        match_log = newlink[1][4]
        #        failed_link(cur, newlink[0], newlink[1], 5, match_log)

        if command == 'link-test':

            # print a summary
            for i, category in enumerate(categories):
        
                print()
                print('Match category ' + str(i) + ': ' + str(len(category)))
                if i != 4 or (i == 4 and len(category) < 20):
                    for match in category:
                        print()
                        prettymatch(match)
        
            print()
            for i, category in enumerate(categories):
        
                print()
                print('Match category ' + str(i) + ': ' + str(len(category)))

        elif command == 'link':

            con.commit()

    con.close()




def make_person_idkey(idkey, given, surname):

    return idkey + '_' + re.sub(r'[^A-Z]', r'', given.strip().upper())[:10].rjust(10, '-') + '_' + re.sub(r'[^A-Z]', r'', surname.strip().upper())[:10].rjust(10, '-')


def most_common(lst):
    return max(set(lst), key=lst.count)


if (command == 'person' or command == 'person-test'):

    con = sqlite3.connect('british-army-officers.db')
    cur = con.cursor()

    cur.execute('CREATE INDEX IF NOT EXISTS idx_link_aidkey ON link(aidkey);')
    cur.execute('CREATE INDEX IF NOT EXISTS idx_link_bidkey ON link(bidkey);')

    if (command == 'person'):

        cur.execute('CREATE TABLE IF NOT EXISTS person(idkey PRIMARY KEY, reg_name, given, surname, row_count);')
        cur.execute('DELETE FROM person;')
        cur.execute('UPDATE row SET person = NULL, link_category = NULL, link_score = NULL, link_log = NULL, linked_to_file = NULL, linked_to_sheet = NULL, linked_to_row = NULL, linked_to_rowid = NULL, row_count = 1;')

    done = set()

    cur.execute('SELECT idkey, reg_name, given, surname FROM row WHERE sheet = "99. All"')

    rows = cur.fetchall()

    for i, row in enumerate(tqdm(rows, total=len(rows), miniters=1, ascii=True, ncols=60)):

        idkey = row[0]

        reg_names = []
        givens = []
        surnames = []

        reg_names.append(row[1])
        givens.append(row[2])
        surnames.append(row[3])

        if idkey not in done:

            found = set()
            found.add(idkey)

            searched = set()
            stilltosearch = set()
            stilltosearch.add(idkey)

            while len(stilltosearch) > 0:

                tosearch = sorted(stilltosearch)[0]

                cur.execute('SELECT aidkey, bidkey, areg_name, breg_name, agiven, bgiven, asurname, bsurname FROM link WHERE link_category < 4 AND (aidkey = ? OR bidkey = ?);', [tosearch, tosearch])

                searched.add(tosearch)

                links = cur.fetchall()

                for link in links:
     
                    found.update([link[0], link[1]])

                    reg_names.append(link[2])
                    reg_names.append(link[3])

                    givens.append(link[4])
                    givens.append(link[5])

                    surnames.append(link[6])
                    surnames.append(link[7])


                stilltosearch = found.difference(searched)

            #if len(found) > 1:
            if True:

                reg_name = most_common(reg_names)
                given = most_common(givens)
                surname = most_common(surnames)

                newpersonidkey = make_person_idkey(idkey, given, surname)

                if (command == 'person-test'):

                    print(newpersonidkey + ' ' + str(len(found)))

                if (command == 'person'):

                    # update person
                    cur.execute('INSERT INTO person VALUES(?, ?, ?, ?, ?);', ([newpersonidkey, reg_name, given, surname, len(found)]))

                foundlist = sorted(found)
                params = sorted(found)
                params.insert(0, len(found))
                params.insert(0, newpersonidkey)

                if (command == 'person'):
                    
                    cur.execute('UPDATE row SET person = ?, row_count = ? WHERE idkey IN (%s);' % ('?,' * len(foundlist))[:-1], params)

            done.update(found)

    if (command == 'person'):

        con.commit()
        con.close()











def find_successor_idkey(idkey, cur):

    idkey65 = '65' + idkey[2:]

    # Does the 65 version of the row exist?
    cur.execute('SELECT idkey FROM row WHERE idkey = ?;', [idkey65])
    res = cur.fetchone()
   
    if (res is None):

        print('65 VERSION OF THE IDKEY DOES NOT EXIST: ' + idkey + ' / ' + idkey65)
        return None


    cur.execute('SELECT idkey, aidkey, bidkey, link_category FROM link WHERE aidkey = ?;', [idkey65])

    existing_links = cur.fetchall()

    if len(existing_links) == 1:

        existing_link = existing_links[0]

        #existing_link_idkey = existing_link[0]
        #existing_aidkey = existing_link[1]
        successor_row_idkey = existing_link[2]
        #existing_link_category = existing_link[3]

        
        # Does the successor row exist?

        cur.execute('SELECT idkey FROM row WHERE idkey = ?;', [successor_row_idkey])
        res = cur.fetchone()

        if (res is None):

            print('SUCCESSOR ROW WAS ALSO AN ORPHAN!')
            return None

        else:

            return successor_row_idkey;

    else:

        print('WRONG NUMBER OF EXISTING LINKS: ' + str(len(existing_links)))

        for existing_link in existing_links:

            print(existing_link)

        return None

    return None



def check_valid_idkeys(cur, aidkey, bidkey):

    # Does the aidkey in this manual link exist?

    successor_aidkey = False

    cur.execute('SELECT idkey FROM row WHERE idkey = ?;', [aidkey])
    res = cur.fetchone()

    if (res is None):

        print('ORPHAN AIDKEY: ' + aidkey)

        successor_idkey = find_successor_idkey(aidkey, cur)
        
        if successor_idkey:

            print('SUCCESSOR IDKEY: ' + successor_idkey)
            aidkey = successor_idkey
            successor_aidkey = True

        else:

            print('COULD NOT FIND SUCCESSOR IDKEY FOR: ' + aidkey)
            aidkey = None;
            exit(1)


    # Does the bidkey in this manual link exist?

    successor_bidkey = False

    cur.execute('SELECT idkey FROM row WHERE idkey = ?;', [bidkey])
    res = cur.fetchone()

    if (res is None):

        print('ORPHAN BIDKEY: ' + bidkey)

        successor_idkey = find_successor_idkey(bidkey, cur)
        
        if successor_idkey:

            print('SUCCESSOR IDKEY: ' + successor_idkey)
            bidkey = successor_idkey
            successor_bidkey = True

        else:

            print('COULD NOT FIND SUCCESSOR IDKEY FOR ' + bidkey)
            bidkey = None
            exit(1)


    return aidkey, bidkey, successor_aidkey, successor_bidkey






def add_link(command, link_idkey, aidkey, bidkey, link_category):

    print('ADD: ', link_idkey, aidkey, bidkey, link_category);

    if (command == 'manual'):

        while True:
            try:
                cur.execute('INSERT INTO link VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?);', ([link_idkey, aidkey, bidkey, link_category, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None]))   
            except Exception as error:
                print("DB Problem:", error, idkey, basename, sheet, row)
                sleep(2)
                continue
            else:
                break




def delete_link(command, link_idkey, aidkey, bidkey, link_category):

    print('DEL: ', link_idkey, aidkey, bidkey, link_category);

    if (command == 'manual'):

        while True:
            try:
                cur.execute('DELETE FROM link WHERE idkey = ? AND aidkey = ? AND bidkey = ? AND link_category = ?;', ([link_idkey, aidkey, bidkey, link_category]))
            except Exception as error:
                print("DB Problem:", error, idkey, basename, sheet, row)
                sleep(2)
                continue
            else:
                break






category_a_sheets = ['10', '11', '12', '14', '15', '16', '17', '22', '23', '24', '25', '26', '27', '28', '32', '55', '56'] # Category A. Should not contain duplicates.
category_c_sheets = ['01', '04', '07', '08', '13', '35'] # Category C. Each person may or may not also be in a Category A Sheet. Used in internal linking stage one.
category_b_sheets = ['02', '03', '05', '06', '09', '34', '36'] # Category B. Each person should definitely also be in a Category A Sheet. Used in internal linking stage two. Should never be able to get into Sheet 99.

if (command == 'manual' or command == 'manual-test'):

    con = sqlite3.connect('british-army-officers.db')
    cur = con.cursor()

    if command == 'manual': create_link_table(cur)

    subcommand = sys.argv[2]

    skip = 0;

    if len(sys.argv) >= 4:

        skip = int(sys.argv[3])
        print('SKIP ' + str(skip))

    end = sys.maxsize;

    if len(sys.argv) == 5:

        end = int(sys.argv[4])
        print('END ' + str(end))

    cur.execute('SELECT id, linkkey, aidkey, bidkey, decision FROM manual;')
    manrows = cur.fetchall()

    for i, mr in enumerate(manrows):

        if i < skip or i > end: continue

        manual_link_id = mr[0]
        manual_link_idkey = mr[1]

        aidkey = mr[2]
        bidkey = mr[3]
        decision = mr[4]

        a99 = aidkey[0:2]
        ayear = aidkey[3:5]
        asheet = aidkey[6:8]
        arow = aidkey[9:15]

        b99 = bidkey[0:2]
        byear = bidkey[3:5]
        bsheet = bidkey[6:8]
        brow = bidkey[9:15]


        # Some manual links we know are bad and are under investigation so we skip them

        if manual_link_id == '300' and manual_link_idkey == '99-49-01-001828_99-48-10-000727': continue
        if manual_link_id == '301' and manual_link_idkey == '99-49-01-001829_99-48-10-000854': continue
        if manual_link_id == '321' and manual_link_idkey == '65-49-01-001864_99-49-10-001063': continue
        if manual_link_id == '3165' and	manual_link_idkey == '99-73-17-000525_99-71-17-000322' : continue
        if manual_link_id == '419' and manual_link_idkey == '99-49-28-001659_99-48-01-000183' : continue

        if manual_link_id == '250' and manual_link_idkey == '99-56-13-000282_99-55-13-000276' : continue
        if manual_link_id == '594' and manual_link_idkey == '99-65-01-000942_99-64-01-001236' : continue
        if manual_link_id == '829' and manual_link_idkey == '99-65-28-002565_99-64-10-000928' : continue
        if manual_link_id == '5113' and manual_link_idkey == '99-67-28-001522_99-66-28-000734' : continue
        if manual_link_id == '5114' and manual_link_idkey == '99-67-28-001541_99-66-28-000753' : continue
        if manual_link_id == '5115' and manual_link_idkey == '99-67-28-001543_99-66-28-000755' : continue
        if manual_link_id == '6612' and manual_link_idkey == '99-64-01-001191_99-63-10-001670' : continue
        if manual_link_id == '6613' and manual_link_idkey == '99-64-01-001203_99-63-10-001528' : continue
        if manual_link_id == '6828' and manual_link_idkey == '99-64-10-008341_99-63-10-007997' : continue
        if manual_link_id == '6829' and manual_link_idkey == '99-64-10-008362_99-63-10-008017' : continue

        #if manual_link_id == '3583' and manual_link_idkey == '99-73-01-001648_99-73-28-003241' : continue
        #if manual_link_id == '3586' and manual_link_idkey == '99-73-08-000052_99-73-14-000032' : continue

        if manual_link_id == '658' and manual_link_idkey == '99-65-10-000185_99-64-10-000191' : continue # MANUAL LINK ALREADY EXISTS FOR aidkey. CONFLICTING MANUAL LINKS?

        if manual_link_id == '4173' and manual_link_idkey == '99-69-01-001651_99-67-01-001706' : continue # MANUAL REJECTION EXPLICITY REJECTS MANUAL LINK.

        if manual_link_id == '472' and manual_link_idkey == '99-65-01-000114_99-64-01-000130' : continue # MANUAL LINK ALREADY EXISTS FOR aidkey. SUCCESSOR aidkey. PROBABLY NEEDS FURTHER INVESTIGATION.
        if manual_link_id == '476' and manual_link_idkey == '99-65-01-000142_99-64-10-000628' : continue # MANUAL LINK ALREADY EXISTS FOR aidkey. SUCCESSOR aidkey. PROBABLY NEEDS FURTHER INVESTIGATION.
        if manual_link_id == '516' and manual_link_idkey == '99-65-01-000451_99-64-01-000476' : continue # MANUAL LINK ALREADY EXISTS FOR aidkey. SUCCESSOR aidkey. PROBABLY NEEDS FURTHER INVESTIGATION.

        # Can we classify this manual link?

        linktype = None

        if (ayear == byear):

            if (asheet in category_c_sheets and bsheet in category_a_sheets):

                linktype = 'stage-one'

        else:

            if (a99 == '99' and b99 == '99' and int(ayear) > int(byear) and (int(ayear) - int(byear)) < 5):
            # We can't be as strict as we would like with the year test because this is archive codes which later on have a bigger gap than one per year eg 1819 -> 1820 is 71 -> 73.

                linktype = 'external'

        if linktype == None:

            print('COULD NOT DETERMINE A LINKTYPE FOR THIS MANUAL LINK:')
            print(mr)
            exit(1)

        if (subcommand == 'pre99-confirm' and linktype == 'pre99' and decision == 'confirm'): # Applies any manual links which change the initial contents of sheet 99 (i.e. violate assumption that Category A sheets do not contain duplicates)

            print()
            print('----------------')
            print(mr)
            print('ACTIONING')
            aidkey, bidkey, successor_aidkey, successor_bidkey = check_valid_idkeys(cur, aidkey, bidkey)
            print(linktype, decision, aidkey, '(SUCCESSOR)' if successor_aidkey else '', bidkey, '(SUCCESSOR)' if successor_bidkey else '')

        if (subcommand == 'pre99-reject' and linktype == 'pre99' and decision == 'reject'): # Removes any links which change the initial contents of sheet 99 (Shold not happen aside from possibly contradicting previous manual pre99 links)

            print()
            print('----------------')
            print(mr)
            print('ACTIONING')
            aidkey, bidkey, successor_aidkey, successor_bidkey = check_valid_idkeys(cur, aidkey, bidkey)
            print(linktype, decision, aidkey, '(SUCCESSOR)' if successor_aidkey else '', bidkey, '(SUCCESSOR)' if successor_bidkey else '')
            print('ACTIONING')

        if (subcommand == 'stage-one-confirm' and linktype == 'stage-one' and decision == 'confirm'): # Links the Category C sheets to Sheet 99. Category C sheets are where each person *may or may not* also be in a Category A Sheet.

            print()
            print('----------------')
            print(mr)
            print('ACTIONING')
            # Stage one links are always from a 65 record to a 99 record, even if the manual link might have been made from a Catecory C record which was later copied into sheet 99.
            aidkey = '65' + aidkey[2:]
            aidkey, bidkey, successor_aidkey, successor_bidkey = check_valid_idkeys(cur, aidkey, bidkey)
            print(linktype, decision, aidkey, '(SUCCESSOR)' if successor_aidkey else '', bidkey, '(SUCCESSOR)' if successor_bidkey else '')

            cur.execute('SELECT idkey, aidkey, bidkey, link_category FROM link WHERE aidkey = ?;', [aidkey])

            existing_links = cur.fetchall()

            print(str(len(existing_links)) + ' EXISTING LINK(S) FOR ' + aidkey + ':')

            link_already_exists = False

            for existing_link in existing_links:

                existing_link_idkey = existing_link[0]
                existing_link_aidkey = existing_link[1]
                existing_link_bidkey = existing_link[2]
                existing_link_category = existing_link[3]

                print('- ', existing_link_aidkey, existing_link_bidkey, existing_link_category);

                if (existing_link_category == 4 or existing_link_category == 7):

                    print('FAILED AUTOMATIC LINK IS BEING REPLACED WITH MANUAL LINK. REMOVE FAILED AUTOMATIC LINK.')
                    delete_link(command, existing_link_idkey, existing_link_aidkey, existing_link_bidkey, existing_link_category)

                elif (existing_link_category == -1):

                    print('MANUAL LINK ALREADY EXISTS FOR aidkey.')

                    if (aidkey == existing_link_aidkey and bidkey == existing_link_bidkey):

                        print('MANUAL LINK ALREADY EXISTS. DO NOTHING.')
                        link_already_exists = True

                    else:

                        if (successor_aidkey):
                    
                            print('SUCCESSOR aidkey. PROBABLY NEEDS FURTHER INVESTIGATION.')
                            exit(1)


                        if (successor_bidkey):
                    
                            print('LINK NEEDS TO BE UPDATED TO SUCCESSOR bidkey. REMOVE AND REPLACE WITH MANUAL LINK.')
                            delete_link(command, existing_link_idkey, existing_link_aidkey, existing_link_bidkey, existing_link_category)

                        else:

                            print('CONFLICTING MANUAL LINKS?')
                            exit(1)

                elif (existing_link_category == 0 or existing_link_category == 1 or existing_link_category == 2 or existing_link_category == 3):

                    print('AUTOMATED LINK ALREADY EXISTS FOR aidkey.')
                    
                    if (aidkey == existing_link_aidkey and bidkey == existing_link_bidkey):

                        print('MANUAL LINK SIMPLY CONFIRMS AUTOMATIC LINK.')
                        print('REMOVE THE AUTOMATIC LINK. THE MANUAL LINK TAKES PRECEDENCE.')
                        delete_link(command, existing_link_idkey, existing_link_aidkey, existing_link_bidkey, existing_link_category)

                    else:

                        print('MANUAL LINK IS TRYING TO LINK THE aidkey TO A DIFFERENT bidkey.')
                        print('REMOVE THE AUTOMATIC LINK. THE MANUAL LINK TAKES PRECEDENCE.')
                        delete_link(command, existing_link_idkey, existing_link_aidkey, existing_link_bidkey, existing_link_category)

                elif (existing_link_category == 6):

                    print('aidkey HAS A MANUAL REJECTION.')

                    if (bidkey == existing_link_bidkey):

                        print('MANUAL REJECTION DIRECTLY CONTRADICTS THE NEW MANUAL LINK')
                        exit(1)

                    else:

                        print('DOES NOT CONTRADICT NEW MANUAL LINK SO LEAVE ALONE.')

                else:

                    print('NOT SURE')
                    exit(1)

            if len(existing_links) != 1:

                print('NUMBER OF EXISTING LINKS FOR ' + aidkey + ' WAS NOT 1.');
                exit(1)

            if link_already_exists == False:

                link_idkey = make_link_idkey(aidkey, bidkey)
                link_category = -1;

                add_link(command, link_idkey, aidkey, bidkey, link_category)



        if (subcommand == 'stage-one-reject' and linktype == 'stage-one' and decision == 'reject'): # Links the Category C sheets to Sheet 99. Category C sheets are where each person *may or may not* also be in a Category A Sheet.

            print()
            print('----------------')
            print(mr)
            print('ACTIONING')
            # Stage one links are always from a 65 record to a 99 record, even if the manual link might have been made from a Catecory C record which was later copied into sheet 99.
            aidkey = '65' + aidkey[2:]
            aidkey, bidkey, successor_aidkey, successor_bidkey = check_valid_idkeys(cur, aidkey, bidkey)
            print(linktype, decision, aidkey, '(SUCCESSOR)' if successor_aidkey else '', bidkey, '(SUCCESSOR)' if successor_bidkey else '')

            cur.execute('SELECT idkey, aidkey, bidkey, link_category FROM link WHERE aidkey = ?;', [aidkey])

            existing_links = cur.fetchall()

            print(str(len(existing_links)) + ' EXISTING LINK(S) FOR ' + aidkey + ':')

            manual_rejection_is_needed = False

            for existing_link in existing_links:

                existing_link_idkey = existing_link[0]
                existing_link_aidkey = existing_link[1]
                existing_link_bidkey = existing_link[2]
                existing_link_category = existing_link[3]

                print('- ', existing_link_aidkey, existing_link_bidkey, existing_link_category);

                if (existing_link_category == 6):

                    print('MANUAL REJECTION ALREADY EXISTS FOR aidkey.')

                    if (bidkey == existing_link_bidkey):

                        print('bidkeys MATCH. THIS MANUAL REJECTION ALREADY EXISTS.')
                        manual_rejection_is_needed = False
                  
                    else:

                        print('MANUAL REJECTION IS OF A DIFFERENT bidkey.')
                        exit(1)

                elif (existing_link_category == 0 or existing_link_category == 1 or existing_link_category == 2 or existing_link_category == 3):

                     print('SUCCESSFUL AUTOMATIC LINK EXISTS FOR aidkey.')

                     if (bidkey == existing_link_bidkey):

                        print('bidkeys MATCH. MANUAL REJECTION EXPLICITY REJECTS THIS AUTOMATIC LINK. REMOVE.')
                        delete_link(command, existing_link_idkey, existing_link_aidkey, existing_link_bidkey, existing_link_category)
                        manual_rejection_is_needed = True

                     else:

                        print('bidkeys DO NOT MATCH. THIS IS A LINK FROM THE aidkey TO SOMEWHERE ELSE. SINCE THIS IS THE ONLY LINK FROM THIS aidkey, THIS MANUAL REJECTION IS NOT NEEDED.')
                        manual_rejection_is_needed = False

                elif (existing_link_category == -1):

                     print('SUCCESSFUL MANUAL LINK EXISTS FOR aidkey.')

                     if (bidkey == existing_link_bidkey):

                        print('bidkeys MATCH. MANUAL REJECTION EXPLICITY REJECTS MANUAL LINK.')
                        exit(1)

                     else:

                        print('bidkeys DO NOT MATCH. THIS IS A LINK FROM THE aidkey TO SOMEWHERE ELSE. SINCE THIS IS THE ONLY LINK FROM THIS aidkey, THIS MANUAL REJECTION IS NOT NEEDED.')
                        manual_rejection_is_needed = False

                elif (existing_link_category == 4 or existing_link_category == 7):

                     print('FAILED AUTOMATIC LINK EXISTS FOR aidkey.')

                     if (bidkey == existing_link_bidkey):

                        print('bidkeys MATCH. MANUAL REJECTION EXPLICITY REJECTS THIS LINK WHICH IS ALREADY FAILED. REMOVE.')
                        delete_link(command, existing_link_idkey, existing_link_aidkey, existing_link_bidkey, existing_link_category)
                        manual_rejection_is_needed = True

                     else:

                        print('bidkeys DO NOT MATCH. THIS IS A LINK FROM THE aidkey TO SOMEWHERE ELSE. SINCE THIS IS THE ONLY LINK FROM THIS aidkey, THIS MANUAL REJECTION IS NOT NEEDED.')
                        manual_rejection_is_needed = False

                else:

                    print('NOT SURE')
                    exit(1)

            if len(existing_links) != 1:

                print('NUMBER OF EXISTING LINKS FOR ' + aidkey + ' WAS NOT 1.');
                exit(1)

            if manual_rejection_is_needed == True:

                link_idkey = make_link_idkey(aidkey, bidkey)
                link_category = 6;

                add_link(command, link_idkey, aidkey, bidkey, link_category)


        if (subcommand == 'stage-two-confirm' and linktype == 'stage-two' and decision == 'confirm'): # Links the Category B sheets to Sheet 99. Category B sheets are where each person should definitely also be in a Category A Sheet.

            print()
            print('----------------')
            print(mr)
            print('ACTIONING')
            aidkey, bidkey, successor_aidkey, successor_bidkey = check_valid_idkeys(cur, aidkey, bidkey)
            print(linktype, decision, aidkey, '(SUCCESSOR)' if successor_aidkey else '', bidkey, '(SUCCESSOR)' if successor_bidkey else '')

        if (subcommand == 'stage-two-reject' and linktype == 'stage-two' and decision == 'reject'): # Links the Category B sheets to Sheet 99. Category B sheets are where each person should definitely also be in a Category A Sheet.

            print()
            print('----------------')
            print(mr)
            print('ACTIONING')
            aidkey, bidkey, successor_aidkey, successor_bidkey = check_valid_idkeys(cur, aidkey, bidkey)
            print(linktype, decision, aidkey, '(SUCCESSOR)' if successor_aidkey else '', bidkey, '(SUCCESSOR)' if successor_bidkey else '')

        if (subcommand == 'external-confirm' and linktype == 'external' and decision == 'confirm'): # Creates links between Sheet 99s and Sheet 99s in previous years only. No other type of external link should exist.

            print()
            print('----------------')
            print(mr)
            print('ACTIONING')
            aidkey, bidkey, successor_aidkey, successor_bidkey = check_valid_idkeys(cur, aidkey, bidkey)
            print(linktype, decision, aidkey, '(SUCCESSOR)' if successor_aidkey else '', bidkey, '(SUCCESSOR)' if successor_bidkey else '')

            cur.execute('SELECT idkey, aidkey, bidkey, link_category FROM link WHERE aidkey = ?;', [aidkey])

            existing_links = cur.fetchall()

            print(str(len(existing_links)) + ' EXISTING LINK(S) FOR ' + aidkey + ':')

            link_already_exists = False

            for existing_link in existing_links:

                existing_link_idkey = existing_link[0]
                existing_link_aidkey = existing_link[1]
                existing_link_bidkey = existing_link[2]
                existing_link_category = existing_link[3]

                print('- ', existing_link_aidkey, existing_link_bidkey, existing_link_category);

                if (existing_link_category == 4 or existing_link_category == 7):

                    print('FAILED AUTOMATIC LINK IS BEING REPLACED WITH MANUAL LINK. REMOVE FAILED AUTOMATIC LINK.')
                    delete_link(command, existing_link_idkey, existing_link_aidkey, existing_link_bidkey, existing_link_category)

                elif (existing_link_category == -1):

                    print('MANUAL LINK ALREADY EXISTS FOR aidkey.')

                    if (aidkey == existing_link_aidkey and bidkey == existing_link_bidkey):

                        print('MANUAL LINK ALREADY EXISTS. DO NOTHING.')
                        link_already_exists = True

                    else:

                        if (successor_aidkey):
                    
                            print('SUCCESSOR aidkey. PROBABLY NEEDS FURTHER INVESTIGATION.')
                            exit(1)


                        if (successor_bidkey):
                    
                            print('LINK NEEDS TO BE UPDATED TO SUCCESSOR bidkey. REMOVE AND REPLACE WITH MANUAL LINK.')
                            delete_link(command, existing_link_idkey, existing_link_aidkey, existing_link_bidkey, existing_link_category)

                        else:

                            print('CONFLICTING MANUAL LINKS?')
                            exit(1)

                elif (existing_link_category == 0 or existing_link_category == 1 or existing_link_category == 2 or existing_link_category == 3):

                    print('AUTOMATED LINK ALREADY EXISTS FOR aidkey.')
                    
                    if (aidkey == existing_link_aidkey and bidkey == existing_link_bidkey):

                        print('MANUAL LINK SIMPLY CONFIRMS AUTOMATIC LINK.')
                        print('REMOVE THE AUTOMATIC LINK. THE MANUAL LINK TAKES PRECEDENCE.')
                        delete_link(command, existing_link_idkey, existing_link_aidkey, existing_link_bidkey, existing_link_category)

                    else:

                        print('MANUAL LINK IS TRYING TO LINK THE aidkey TO A DIFFERENT bidkey.')
                        print('REMOVE THE AUTOMATIC LINK. THE MANUAL LINK TAKES PRECEDENCE.')
                        delete_link(command, existing_link_idkey, existing_link_aidkey, existing_link_bidkey, existing_link_category)

                elif (existing_link_category == 6):

                    print('aidkey HAS A MANUAL REJECTION.')

                    if (bidkey == existing_link_bidkey):

                        print('MANUAL REJECTION DIRECTLY CONTRADICTS THE NEW MANUAL LINK')
                        exit(1)

                    else:

                        print('DOES NOT CONTRADICT NEW MANUAL LINK SO LEAVE ALONE.')

                else:

                    print('NOT SURE')
                    exit(1)

            if len(existing_links) != 1:

                print('NUMBER OF EXISTING LINKS FOR ' + aidkey + ' WAS NOT 1.');
                exit(1)

            if link_already_exists == False:

                link_idkey = make_link_idkey(aidkey, bidkey)
                link_category = -1;

                add_link(command, link_idkey, aidkey, bidkey, link_category)


        if (subcommand == 'external-reject' and linktype == 'external' and decision == 'reject'): # Creates links between Sheet 99s and Sheet 99s in previous years only. No other type of external link should exist.

            print()
            print('----------------')
            print(mr)
            print('ACTIONING')
            aidkey, bidkey, successor_aidkey, successor_bidkey = check_valid_idkeys(cur, aidkey, bidkey)
            print(linktype, decision, aidkey, '(SUCCESSOR)' if successor_aidkey else '', bidkey, '(SUCCESSOR)' if successor_bidkey else '')

            cur.execute('SELECT idkey, aidkey, bidkey, link_category FROM link WHERE aidkey = ?;', [aidkey])

            existing_links = cur.fetchall()

            print(str(len(existing_links)) + ' EXISTING LINK(S) FOR ' + aidkey + ':')

            manual_rejection_is_needed = False

            for existing_link in existing_links:

                existing_link_idkey = existing_link[0]
                existing_link_aidkey = existing_link[1]
                existing_link_bidkey = existing_link[2]
                existing_link_category = existing_link[3]

                print('- ', existing_link_aidkey, existing_link_bidkey, existing_link_category);

                if (existing_link_category == 6):

                    print('MANUAL REJECTION ALREADY EXISTS FOR aidkey.')

                    if (bidkey == existing_link_bidkey):

                        print('bidkeys MATCH. THIS MANUAL REJECTION ALREADY EXISTS.')
                        manual_rejection_is_needed = False
                  
                    else:

                        print('MANUAL REJECTION IS OF A DIFFERENT bidkey.')
                        exit(1)

                elif (existing_link_category == 0 or existing_link_category == 1 or existing_link_category == 2 or existing_link_category == 3):

                     print('SUCCESSFUL AUTOMATIC LINK EXISTS FOR aidkey.')

                     if (bidkey == existing_link_bidkey):

                        print('bidkeys MATCH. MANUAL REJECTION EXPLICITY REJECTS THIS AUTOMATIC LINK. REMOVE.')
                        delete_link(command, existing_link_idkey, existing_link_aidkey, existing_link_bidkey, existing_link_category)
                        manual_rejection_is_needed = True

                     else:

                        print('bidkeys DO NOT MATCH. THIS IS A LINK FROM THE aidkey TO SOMEWHERE ELSE. SINCE THIS IS THE ONLY LINK FROM THIS aidkey, THIS MANUAL REJECTION IS NOT NEEDED.')
                        manual_rejection_is_needed = False

                elif (existing_link_category == -1):

                     print('SUCCESSFUL MANUAL LINK EXISTS FOR aidkey.')

                     if (bidkey == existing_link_bidkey):

                        print('bidkeys MATCH. MANUAL REJECTION EXPLICITY REJECTS MANUAL LINK.')
                        exit(1)

                     else:

                        print('bidkeys DO NOT MATCH. THIS IS A LINK FROM THE aidkey TO SOMEWHERE ELSE. SINCE THIS IS THE ONLY LINK FROM THIS aidkey, THIS MANUAL REJECTION IS NOT NEEDED.')
                        manual_rejection_is_needed = False

                elif (existing_link_category == 4 or existing_link_category == 7):

                     print('FAILED AUTOMATIC LINK EXISTS FOR aidkey.')

                     if (bidkey == existing_link_bidkey):

                        print('bidkeys MATCH. MANUAL REJECTION EXPLICITY REJECTS THIS LINK WHICH IS ALREADY FAILED. REMOVE.')
                        delete_link(command, existing_link_idkey, existing_link_aidkey, existing_link_bidkey, existing_link_category)
                        manual_rejection_is_needed = True

                     else:

                        print('bidkeys DO NOT MATCH. THIS IS A LINK FROM THE aidkey TO SOMEWHERE ELSE. SINCE THIS IS THE ONLY LINK FROM THIS aidkey, THIS MANUAL REJECTION IS NOT NEEDED.')
                        manual_rejection_is_needed = False

                else:

                    print('NOT SURE')
                    exit(1)

            if len(existing_links) != 1:

                print('NUMBER OF EXISTING LINKS FOR ' + aidkey + ' WAS NOT 1.');
                exit(1)

            if manual_rejection_is_needed == True:

                link_idkey = make_link_idkey(aidkey, bidkey)
                link_category = 6;

                add_link(command, link_idkey, aidkey, bidkey, link_category)




    print()
    print('ALL MANUAL LINKS READ WITH NO OBVIOUS PROBLEMS.')
    print()

    if (command == 'manual'):

        con.commit()
        con.close()

if (command == 'internal-link-redirect' or command == 'internal-link-redirect-test'):

    con = sqlite3.connect('british-army-officers.db')
    cur = con.cursor()

    skip = 0;

    if len(sys.argv) >= 3:

        skip = int(sys.argv[2])
        print('SKIP ' + str(skip))

    end = sys.maxsize;

    if len(sys.argv) == 4:

        end = int(sys.argv[3])
        print('END ' + str(end))

    cur.execute('SELECT idkey, aidkey, bidkey, link_category FROM link;')
    rows = cur.fetchall()

    bad_link_count = 0;
    no_v99_count = 0;
    v99_count = 0;
    successor_count = 0;
    no_successor_count = 0;

    for i, row in enumerate(tqdm(rows, total=len(rows), miniters=1, ascii=True, ncols=60)):

        if i < skip or i > end: continue

        idkey = row[0]
        aidkey = row[1]
        bidkey = row[2]
        link_category = row[3]

        a99 = aidkey[0:2]
        ayear = aidkey[3:5]
        asheet = aidkey[6:8]
        arow = aidkey[9:15]

        b99 = bidkey[0:2]
        byear = bidkey[3:5]
        bsheet = bidkey[6:8]
        brow = bidkey[9:15]

        if (ayear == byear and a99 == '65' and b99 == '65'):

                #print('bad link: ' + aidkey + ' --> ' + bidkey);

                bad_link_count = bad_link_count + 1;

                bidkey99 = '99' + bidkey[2:]

                cur.execute('SELECT idkey FROM row WHERE idkey = ?;', [bidkey99])
                res = cur.fetchone()
   
                if (res is None):

                    no_v99_count = no_v99_count + 1;

                    print('No 99 version: ' + bidkey)

                    # If there is no 99 version, did the bidkey get linked to something else?

                    cur.execute('SELECT idkey, aidkey, bidkey, link_category FROM link WHERE aidkey = ?;', [bidkey])

                    existing_links = cur.fetchall()

                    if len(existing_links) == 1:

                        successor_count = successor_count + 1;

                        existing_link = existing_links[0]
                        successor_idkey = existing_link[2]

                        print('Successor idkey: ' + successor_idkey)

                        print("UPDATE link set bidkey = '" + successor_idkey + "', idkey = '" + (aidkey + '_' + successor_idkey) + "' WHERE idkey = '" + idkey + "';");

                        if (command == 'internal-link-redirect'):

                            while True:
                                try:
                                    cur.execute('UPDATE link set bidkey = ?, idkey = ? WHERE idkey = ?;', ([successor_idkey, (aidkey + '_' + successor_idkey), idkey]))   
                                except Exception as error:
                                    print("DB Problem:", error, idkey, basename, sheet, row)
                                    sleep(2)
                                    continue
                                else:
                                    break

                    else:

                        print('No successor idkey.')
                        no_successor_count = no_successor_count + 1;


                else:

                    v99_count = v99_count + 1;

                    #print('99 version.')

                    print("UPDATE link set bidkey = '" + bidkey99 + "', idkey = '" + (aidkey + '_' + bidkey99) + "' WHERE idkey = '" + idkey + "';");

                    if (command == 'internal-link-redirect'):

                        while True:
                            try:
                                cur.execute('UPDATE link set bidkey = ?, idkey = ? WHERE idkey = ?;', ([bidkey99, (aidkey + '_' + bidkey99), idkey]))   
                            except Exception as error:
                                print("DB Problem:", error, idkey, basename, sheet, row)
                                sleep(2)
                                continue
                            else:
                                break


    print('     bad_link_count: ' + str(bad_link_count))
    print('   99_version_count: ' + str(v99_count))
    print('no_99_version_count: ' + str(no_v99_count))
    print('    successor_count: ' + str(successor_count))
    print(' no_successor_count: ' + str(no_successor_count))

    if (command == 'internal-link-redirect'):

        con.commit()
        con.close()

# -1 Manually Confirmed
#  0 High
#  1 Good
#  2 Fair
#  3 Low
#  4 Failed
#  5 New Appointment
#  6 Manually Rejected
#  7 Better Link Found

