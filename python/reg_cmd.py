#!/usr/bin/python3

import sys
import sqlite3

from reg import reg_regiment, reg_rank, rank_lvl, reg_date, reg_name, reg_date2

#import re
from openpyxl import Workbook, load_workbook
#from datetime import datetime
#import traceback
from tqdm import tqdm
#from nameparser import HumanName
#from reg_number import reg_number

from match import rank_similarity

units = []

#uwb = load_workbook(filename = 'data/lookups/Units.xlsx', read_only=True)
#uws = uwb['Units']

#for xlrow in uws.iter_rows():
#    units.append(xlrow[0].value)

#uwb = load_workbook(filename = 'data/lookups/UnitsNew.xlsx', read_only=True)
#uws = uwb['Sheet1']

#for xlrow in uws.iter_rows(min_row=2):
#    units.append(xlrow[1].value)

uwb = load_workbook(filename = 'data/lookups/UnitsJuly24.xlsx', read_only=True)
uws = uwb['Standard titles']

for xlrow in uws.iter_rows(min_row=2):
    units.append(xlrow[1].value)

#print(units)
#exit()

supp_units = []
#supp_units.append('Foot')
#supp_units.append('')
#supp_units.append('')
#supp_units.append('')
#supp_units.append('')
#supp_units.append('')
#supp_units.append('')
#supp_units.append('')
#supp_units.append('')
#supp_units.append('')
#supp_units.append('')
#supp_units.append('')

command = sys.argv[1]

if (command == 'rank-similarity'):

       
    ranka = sys.argv[2]
    rankb = sys.argv[3]

    areg = reg_rank(ranka)
    alvl = rank_lvl(areg)

    breg = reg_rank(rankb)
    blvl = rank_lvl(breg)

    similarity = rank_similarity(areg, breg, alvl, blvl, 1799, 1800);

    print('    orig a: **' + ranka + '**')
    print('     reg a: **' + areg + '**')
    print('     lvl a: **' + str(alvl) + '**')
    print()
    print('    orig b: **' + rankb + '**')
    print('     reg b: **' + breg + '**')
    print('     lvl b: **' + str(blvl) + '**')
    print()
    print('similarity: **' + str(similarity) + '**')
    exit()




if (command == 'regiment'):

    if len(sys.argv) == 4:
            
        testregiment = sys.argv[2]
        testyear = sys.argv[3]
        rregiment = reg_regiment(testregiment, testyear)
        print('            original: **' + testregiment + '**')
        print('                year: **' + testyear + '**')
        print('                 reg: **' + rregiment + '**')
        print('     recognised unit: **' + str(rregiment in units) + '**')
        print('recognised supp_unit: **' + str(rregiment in supp_units) + '**')
        exit()

    if len(sys.argv) == 3:
            
        testregiment = sys.argv[2]
        rregiment = reg_regiment(testregiment, None)
        print('            original: **' + testregiment + '**')
        print('                 reg: **' + rregiment + '**')
        print('     recognised unit: **' + str(rregiment in units) + '**')
        print('recognised supp_unit: **' + str(rregiment in supp_units) + '**')
        exit()

    con = sqlite3.connect('british-army-officers.db')
    cur = con.cursor()
    cur.execute('SELECT regiment, count(*), fileyear FROM row GROUP BY regiment;')
    rows = cur.fetchall()
    con.close()

    terms = {}

    for i, row in enumerate(rows):

        term = row[0]
        count = row[1]
        fileyear = row[2]

        if isinstance(term, str):

            reg = reg_regiment(term, fileyear)

            if reg in terms: count = terms[reg][1] + count

            terms[reg] = (term, count)

    print()

    #terms2 = {k: v for k, v in sorted(terms.items(), key=lambda item: item[1][1])} # Order by frequency low first
    terms2 = {k: v for k, v in sorted(terms.items(), reverse=True, key=lambda item: item[1][1])} # Order by frequency high first

    for i, key in enumerate(terms2):
        if not key in units:
            #print('{:4.4}'.format(str(i)) + ' ' + '{:80.80}'.format('**' + key + '**') + ' ' + '{:10.10}'.format(str(key in units)) + ' ' + '{:10.10}'.format(str(key in supp_units)) + ' ' + '{:8.8}'.format(str(terms[key][1])) + ' ' + '{:100.100}'.format('**' + terms[key][0] + '**'))
            print('{:80.80}'.format('**' + key + '**') + ' ' + '{:8.8}'.format(str(terms[key][1])) + ' ' + '{:100.100}'.format('**' + terms[key][0] + '**'))

    print()
    print(str(len(rows)) + ' --> ' + str(len(terms)))
    print()




if (command == 'rank'):

    if len(sys.argv) == 3:
            
        testrank = sys.argv[2]
        rrank = reg_rank(testrank)
        rlvl = rank_lvl(rrank)
        print('   original: **' + testrank + '**')
        print('        reg: **' + rrank + '**')
        print('   rank_lvl: **' + str(rlvl) + '**')
        exit()

    con = sqlite3.connect('british-army-officers.db')
    cur = con.cursor()
    cur.execute('SELECT rank, count(*) FROM row GROUP BY rank;')
    rows = cur.fetchall()
    con.close()

    terms = {}

    for i, row in enumerate(rows):

        term = row[0]
        count = row[1]

        # filter out unrecoverable ranks

        if term == '`': term = None

        if term != None:

            reg = reg_rank(term)

            if reg in terms: count = terms[reg][1] + count

            terms[reg] = (term, count)

    print()

    terms2 = {k: v for k, v in sorted(terms.items(), key=lambda item: item[1][1])} # Order by frequency

    for i, key in enumerate(terms2):
        if rank_lvl(key) == None:
            print('{:4.4}'.format(str(i)) + ' ' + '{:60.60}'.format('**' + key + '**') + ' ' + '{:10.10}'.format(str(rank_lvl(key))) + ' ' + '{:8.8}'.format(str(terms[key][1])) + ' ' + '{:100.100}'.format('**' + terms[key][0] + '**'))

    print()
    print(str(len(rows)) + ' --> ' + str(len(terms)))
    print()




if (command == 'date'):

    if len(sys.argv) == 3:
            
        testdate = sys.argv[2]
        rdate = reg_date(testdate)
        print('   original: **' + testdate + '**')
        print('        reg: **' + str(rdate) + '**')
        exit()

    con = sqlite3.connect('british-army-officers.db')
    cur = con.cursor()
    cur.execute('SELECT date FROM row GROUP BY date;')
    rows = cur.fetchall()
    con.close()

    terms = {}
    errors = 0

    for i, row in enumerate(rows):

        term = row[0]

        # filter out unrecoverable dates

        if term == '#? July 1785': term = None
        if term == '0 March 1789': term = None
        if term == '09 February 17891': term = None
        if term == '17 September 1773 19 March 1783': term = None
        if term == '18 February 178': term = None
        if term == '19 March 178.': term = None
        if term == '2023-10-17 00:00:00': term = None
        if term == '21 May 178': term = None
        if term == '30 April.': term = None
        if term == '39 July 1791': term = None
        if term == '5v February 1787': term = None
        if term == '82 February 1791': term = None
        if term == '8bNovember 1778': term = None
        if term == 'April 1776': term = None

        if term != None:

            reg = reg_date(term)

            if reg == None:

                print('Could not parse into date: **' + str(term) + '**')
                errors += 1

            terms[reg] = term

    print()
    print(str(errors) + ' dates did not parse.')
    print()


if (command == 'date2'):

    con = sqlite3.connect('british-army-officers.db')
    cur = con.cursor()
    cur.execute('SELECT date2 FROM row GROUP BY date2;')
    rows = cur.fetchall()
    con.close()

    terms = {}
    errors = 0

    for i, row in enumerate(rows):

        date2 = row[0]

        reg_date2_value, reg_date2rank_value, reg_date2lvl = reg_date2(date2)

        if reg_date2_value is not None or reg_date2rank_value is not None:

            pass
            #print('PASS: **' + str(date2) + '** **' + reg_date2rank_value + '** **' + str(reg_date2lvl) + '** **' + str(reg_date2_value) + '**')

        else:

            print('FAIL: **' +str(date2) + '**')
            errors += 1

        #if isinstance(term, str):

            #print(term)

                    #exit()

        # filter out unrecoverable dates

        #if term == '#? July 1785': term = None
        #if term == '0 March 1789': term = None
        #if term == '09 February 17891': term = None
        #if term == '17 September 1773 19 March 1783': term = None
        #if term == '18 February 178': term = None
        #if term == '19 March 178.': term = None
        #if term == '2023-10-17 00:00:00': term = None
        #if term == '21 May 178': term = None
        #if term == '30 April.': term = None
        #if term == '39 July 1791': term = None
        #if term == '5v February 1787': term = None
        #if term == '82 February 1791': term = None
        #if term == '8bNovember 1778': term = None
        #if term == 'April 1776': term = None

        #if term != None:

        #    reg = reg_date(term)

        #    if reg == None:

        #        print('Could not parse into date: **' + str(term) + '**')
        #        errors += 1

        #    terms[reg] = term

    #print()
    print(str(errors) + ' date2s did not parse.')
    #print()





if (command == 'name'):

    if len(sys.argv) == 3:
        
        testname = sys.argv[2]
        reg, surname, given, middlenames, title, namesuffix, nickname, name_rank = reg_name(testname)
        print('   original: **' + testname + '**')
        print('        reg: **' + reg + '**')
        print()
        print('      title: **' + title + '**')
        print('      given: **' + given + '**')
        print('middlenames: **' + middlenames + '**')
        print('    surname: **' + surname + '**')
        print(' namesuffix: **' + namesuffix + '**')
        print('   nickname: **' + nickname + '**')
        print('  name_rank: **' + name_rank + '**')
        exit()


    con = sqlite3.connect('british-army-officers.db')
    cur = con.cursor()
    cur.execute('SELECT name, count(*) FROM row GROUP BY name;')
    rows = cur.fetchall()
    con.close()

    reg_names = {}
    givens = {}
    surnames = {}

    for i, row in enumerate(rows):

        term = row[0]

        if True:
        #if 'Lord' in term:
        #if 'K.' in term:

            count = row[1]

            if term != None:

                reg, surname, given, middlenames, title, namesuffix, nickname, name_rank = reg_name(term)

                if reg in reg_names: count = reg_names[reg][1] + count
                reg_names[reg] = (term, count)

                if given in givens: count = givens[given][1] + count
                givens[given] = (term, count)

                if surname in surnames: count = surnames[surname][1] + count
                surnames[surname] = (term, count)

    print()

    #for i, key in enumerate(sorted(reg_names)):
    #    reg_name_value, surname, given, middlenames, title, namesuffix, nickname = reg_name(reg_names[key][0])
    #    print('{:4.4}'.format(str(i)) + ' ' + '{:50.50}'.format('**' + key + '**') + ' ' + '{:8.8}'.format(str(reg_names[key][1])) + ' ' + '{:40.40}'.format('**' + reg_names[key][0] + '**') + ' ' + '{:40.40}'.format('**' + reg_name_value + '**') + ' ' + '{:30.30}'.format('**' + given + '**') + ' ' + '{:30.30}'.format('**' + surname + '**'))

    #for i, key in enumerate(sorted(surnames)):
    #    reg_name_value, surname, given, middlenames, title, namesuffix, nickname = reg_name(surnames[key][0])
    #    print('{:4.4}'.format(str(i)) + ' ' + '{:40.40}'.format('**' + key + '**') + ' ' + '{:8.8}'.format(str(surnames[key][1])) + ' ' + '{:40.40}'.format('**' + surnames[key][0] + '**') + ' ' + '{:40.40}'.format('**' + reg_name_value + '**') + ' ' + '{:30.30}'.format('**' + given + '**'))

    for i, key in enumerate(sorted(givens)):
        reg_name_value, surname, given, middlenames, title, namesuffix, nickname, name_rank = reg_name(givens[key][0])
        print('{:4.4}'.format(str(i)) + ' ' + '{:40.40}'.format('**' + key + '**') + ' ' + '{:8.8}'.format(str(givens[key][1])) + ' ' + '{:40.40}'.format('**' + givens[key][0] + '**') + ' ' + '{:40.40}'.format('**' + reg_name_value + '**') + ' ' + '{:40.40}'.format('**' + surname + '**'))

    print()
    print('Names: ')
    print(str(len(rows)) + ' --> ' + str(len(reg_names)))
    print()
    print('Surnames: ')
    print(str(len(surnames)))
    print()
    print('Givens: ')
    print(str(len(givens)))
    print()





if (command == 're-reg'):

    con = sqlite3.connect('british-army-officers.db')
    cur = con.cursor()
    cur.execute('SELECT rowid, regiment, rank, date, name, date2, rank2, fileyear FROM row;')
    rows = cur.fetchall()

    for i, row in enumerate(tqdm(rows, total=len(rows), miniters=10, ascii=True, ncols=60)):

        reg_name_value, surname, given, middlenames, title, namesuffix, nickname, name_rank = reg_name(row[4])

        name_rank_lvl = rank_lvl(name_rank)

        reg_rank_value = reg_rank(row[2])

        reg_date2_value, reg_date2rank_value, reg_date2lvl_value = reg_date2(row[5])

        reg_rank2_value = reg_rank(row[6])

        cur.execute('UPDATE row SET reg_regiment = ?, reg_rank = ?, rank_lvl = ?, reg_date = ?, reg_name = ?, surname = ?, given = ?, middlenames = ?, title = ?, namesuffix = ?, nickname = ?, reg_date2 = ?, reg_date2rank = ?, reg_date2lvl = ?, reg_rank2 = ?, rank2_lvl = ?, name_rank = ?, name_rank_lvl = ? WHERE rowid = ?;', ([reg_regiment(row[1], row[7]), reg_rank_value, rank_lvl(reg_rank_value), reg_date(row[3]), reg_name_value, surname, given, middlenames, title, namesuffix, nickname, reg_date2_value, reg_date2rank_value, reg_date2lvl_value, reg_rank2_value, rank_lvl(reg_rank2_value), name_rank, name_rank_lvl, row[0]]))

    con.commit()
    con.close()


if (command == 'common'):

    con = sqlite3.connect('british-army-officers.db')
    cur = con.cursor()

    cur.execute('SELECT given, COUNT(given) AS common FROM row GROUP BY given ORDER BY common DESC LIMIT 1;')
    mx = cur.fetchone()[1]
    cur.execute('DROP TABLE IF EXISTS common_given;')
    cur.execute('CREATE TABLE common_given AS SELECT given AS term, cast(COUNT(given) AS FLOAT)/? AS common FROM row GROUP BY given ORDER BY common DESC;', [mx])
    cur.execute('CREATE INDEX IF NOT EXISTS idx_common_given_term ON common_given(term);')
    cur.execute("SELECT term, common FROM common_given WHERE term = 'John';")
    print(cur.fetchone())

    cur.execute('SELECT surname, COUNT(surname) AS common FROM row GROUP BY surname ORDER BY common DESC LIMIT 1;')
    mx = cur.fetchone()[1]
    cur.execute('DROP TABLE IF EXISTS common_surname;')
    cur.execute('CREATE TABLE common_surname AS SELECT surname AS term, cast(COUNT(surname) AS FLOAT)/? AS common FROM row GROUP BY surname ORDER BY common DESC;', [mx])
    cur.execute('CREATE INDEX IF NOT EXISTS idx_common_surname_term ON common_surname(term);')
    cur.execute("SELECT term, common FROM common_surname WHERE term = 'Campbell';")
    print(cur.fetchone())

    con.commit()
    con.close()
        
